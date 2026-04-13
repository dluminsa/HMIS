import os
import io
import time
import traceback
import datetime as dt
from datetime import datetime
from pathlib import Path

import pandas as pd
import numpy as np
import streamlit as st
import gspread
from openpyxl import load_workbook
from google.oauth2.service_account import Credentials
from streamlit_gsheets import GSheetsConnection



#Clear cache at the very start of the app
st.cache_data.clear()
st.cache_resource.clear()

def extract():
    cola,colb,colc = st.columns([1,3,1])
    colb.subheader('106a PIVOT TABLES')   
    today = datetime.now()
    todayd = today.strftime("%Y-%m-%d")# %H:%M")
    week = today.strftime("%V")
    wk = int(week) + 13
    # wk = int(week) - 39 # USE ONLY IN Q4
    thisweek = int(week)-1
    woke = wk-2
    cola,colb = st.columns(2)
    cola.write(f"**DATE TODAY:    {todayd}**")
    colb.write(f"**CURRENT SURGE WEEK:    {wk}**")

    file = st.file_uploader("Upload your EMR extract here", type=['xlsx']) 
    if file is not None:   
        if 'fd' not in st.session_state:
            fileN = file.name
            name = os.path.basename(fileN).split('.')[0]
            st.session_state.fd = name
        else:
            pass
    else:
        pass
    if file is not None: 
       fileN = file.name
       namey = os.path.basename(fileN).split('.')[0]
       if str(namey) != str(st.session_state.fd):
                #st.info(f'DATA FOR {facy} NOT SUBMITTED')
                st.session_state.submited = False
                st.cache_data.clear()
                st.session_state.fd = namey
                st.cache_resource.clear()
                st.session_state.submited =False
                st.session_state.df = None
                st.session_state.reader =False#
                time.sleep(1)
                        
    if 'submited' not in st.session_state:
        st.session_state.submited =False
    if 'df' not in st.session_state:
        st.session_state.df = None
    # if 'fac' not in st.session_state:
   # st.session_state.fac = None
    if 'reader' not in st.session_state:
        st.session_state.reader =False#
    #ext = None
    if file is not None and not st.session_state.reader:
        # Get the file name
        fileN = file.name
        ext = os.path.basename(fileN).split('.')[1]
        # if ext == 'xlsx.xlsx':
        #        ext = 'xlsx'
    #df = None
    if file is not None and not st.session_state.reader:
        wb = load_workbook(file)
        sheets = wb.sheetnames
        if len(sheets)>1:
            st. warning('THIS EXTRACT HAS MULTIPLE SHEETS, I CAN NOT TELL WHICH ONE TO READ')
            time.sleep(3)
            st.info('DELETE ALL THE OTHER SHEETS AND REMAIN WITH ONE THAT HAS THE EVER ENROLLED')
            st.stop()
        else:
            pass

    if file is not None and not st.session_state.reader:
                    st.session_state.df = pd.read_excel(file)
                    df = st.session_state.df
                    st.write('Excel accepted, summaries and linelists below will be for this excel')
                    st.write('To change this excel or to upload another excel, first refresh the page')
    #if file is not None and not st.session_state.rea
                    df = df.rename(columns= {'ART  ':'ART',  'AS  ':'AS', 'RD  ':'RD', 'RD1  ':'RD1', 'RD2  ':'RD2', 'VD  ':'VD', 'FE  ':'FE', 'LD  ': 'LD', 'ARVD  ': 'ARVD',
        'TI  ': 'TI', 'TO  ':'TO', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD', 'NTO  ': 'NTO'})#, 'TPT ': 'TPT'})
                    df = df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'RD ':'RD', 'RD1 ':'RD1', 'RD2 ':'RD2', 'VD ':'VD', 'FE ':'FE', 'LD ': 'LD', 'ARVD ': 'ARVD',
                            'TI ': 'TI', 'TO ':'TO', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD',  'NTO ': 'NTO'})#, 'TPT  ': 'TPT'})
                    columns = ['ART','AG', 'GD','AS', 'VD', 'RD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'ARVD', 'NTO', 'DSD','ARVL','CD']
                    cols = df.columns.to_list()
                    needed = set(columns)
                    there = set(cols)
                    missing = needed - there
                    missing = list(missing)
                    if not all(column in cols for column in columns):
                        missing_columns = [column for column in columns if column not in cols]
                        for column in missing_columns:
                            st.markdown(f' **ERROR !!! MISSING COLUMN(S): {missing}**')
                            st.markdown('**First rename all the columns as guided above**')
                            st.stop()
                    st.session_state.reader= True
    if st.session_state.reader:
                          # Convert 'ART' column to string and create 'ART' column with numeric part to remove blanks
                        st.session_state.df = st.session_state.df.rename(columns= {'ART  ':'ART', 'AS  ':'AS', 'RD  ':'RD', 'RD1  ':'RD1', 'RD2  ':'RD2', 'VD  ':'VD', 'FE  ':'FE', 'LD  ': 'LD', 'ARVD  ': 'ARVD',
        'TI  ': 'TI', 'TO  ':'TO', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD',  'NTO  ': 'NTO'})#, 'TPT ': 'TPT'})
                        st.session_state.df = st.session_state.df.rename(columns= {'ART ':'ART',  'AS ':'AS', 'RD ':'RD', 'RD1 ':'RD1', 'RD2 ':'RD2', 'VD ':'VD', 'FE ':'FE', 'LD ': 'LD', 'ARVD ': 'ARVD',
                            'TI ': 'TI', 'TO ':'TO', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD',  'NTO ': 'NTO'})#, 'TPT  ': 'TPT'})
                        df = st.session_state.df.copy()
                        twagala = df.columns.to_list()
                        uploaded = st.session_state.df[['ART','AS', 'AG','VD', 'RD','GD','TO', 'TI', 'DD', 'FE','LD', 'RD1', 'RD2', 'ARVD', 'NTO']].copy()#,'TPT']].copy()
                        df['ART'] = df['ART'].astype(str)
                        df['A'] = df['ART'].str.replace('[^0-9]', '', regex=True)
                        df['A'] = pd.to_numeric(df['A'], errors= 'coerce')
                        df = df[df['A']>0].copy()
                        #df.dropna(subset='ART', inplace=True)                  
                        # df[['AS', 'RD', 'VD','TO','TI','TPT']] = df[['AS', 'RD', 'VD','TO','TI', 'TPT']].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        if df['TI'].str.contains('YES').any():
                            st.write("You may be using the Transfer in column instead of the Transfer_in Obs date column")
                            st.stop()

                        testrt = df.copy()
                        testra = testrt[~testrt['RD1'].isnull()].copy()
                        if testra.shape[0]<10:
                            st.warning('RD1 is empty, use the correct Return Visit Date1, it can not be blank')
                            testra =df[['ART', 'RD', 'RD1']].copy()
                            st.write(testra.head(5))
                            st.stop()
                            
                        testrb = testrt[~testrt['RD2'].isnull()].copy()
                        if testrb.shape[0]<10:
                            st.warning('RD2 is empty, use the correct Return Visit Date2, it can not be blank')
                            testrb =df[['ART', 'RD', 'RD2']].copy()
                            st.write(testrb.head(5))
                            st.stop()
                        testrc = df[~testrt['VD'].isnull()].copy()    
                        if testrc.shape[0]<10:
                            st.warning('VD is empty, use the correct HIV Viral Load Date, it can not be blank')
                            testrc =df[['ART', 'RD', 'VD']].copy()
                            st.write(testrc.head(5))
                            st.stop()
                        testrd = df[~testrt['LD'].isnull()].copy()    
                        if testrd.shape[0]<10:
                            st.warning('LD is empty, use the correct Last Encouter Date, it can not be blank')
                            testrd =df[['ART', 'RD', 'LD']].copy()
                            st.write(testrd.head(5))
                            st.stop()
                        testre = df[~testrt['AS'].isnull()].copy()    
                        if testre.shape[0]<10:
                            st.warning('AS is empty, use the correct Art Start Date, it can not be blank')
                            testre =df[['ART', 'RD', 'AS']].copy()
                            st.write(testre.head(5))
                            st.stop()
                        df['AS'] = df['AS'].astype(str)
                        df['ARVD'] = df['ARVD'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)                 
                        y = pd.DataFrame({'ART' :['2','3','4','5'], 'TI':['1-1-1',1,'1/1/1','3 8 2001'], 'RD':['1-1-1',1,'1/1/1','3 8 2001'],'DD':['1-1-1',1,'1/1/1','3 8 2001'], 
                                        'TO':['1-1-1',1,'1/1/1','3 8 2001'], 'AS':['1-1-1',1,'1/1/1','3 8 2001'], 'VD':['1-1-1',1,'1/1/1','3 8 2001'],'RD1':['1-1-1',1,'1/1/1','3 8 2001'],
                                        'RD2':['1-1-1',1,'1/1/1','3 8 2001'],'ARVD':['1-1-1',1,'1/1/1','3 8 2001'],
                                        'LD':['1-1-1',1,'1/1/1','3 8 2001'],'FE':['1-1-1',1,'1/1/1','3 8 2001']})                        
                        df['AS'] = df['AS'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)            
                        df['AS'] = df['AS'].str.replace('00:00:00', '', regex=True)
                        df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
                        df['RD1'] = df['RD1'].str.replace('00:00:00', '', regex=True)
                        df['RD2'] = df['RD2'].str.replace('00:00:00', '', regex=True)
                        df['TI'] = df['TI'].str.replace('00:00:00', '', regex=True)
                        df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
                        df['VD'] = df['VD'].str.replace('00:00:00', '', regex=True)
                        df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
                        df['LD'] = df['LD'].str.replace('00:00:00', '', regex=True)
                        df['FE'] = df['FE'].str.replace('00:00:00', '', regex=True) 
                        df["TI"] = df["TI"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["DD"] = df["DD"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["RD"] = df["RD"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["RD1"] = df["RD1"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["RD2"] = df["RD2"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["VD"] = df["VD"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["TO"] = df["TO"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["AS"] = df["AS"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["TI"] = df["TI"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["DD"] = df["DD"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["RD"] = df["RD"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["RD1"] = df["RD1"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["RD2"] = df["RD2"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["VD"] = df["VD"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["TO"] = df["TO"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["AS"] = df["AS"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df = pd.concat([df,y])
                        df = df.copy()
                        df['AS'] = df['AS'].astype(str) ###
                        df['RD'] = df['RD'].astype(str) ###
                        df['RD1'] = df['RD1'].astype(str)##
                        df['RD2'] = df['RD2'].astype(str)##
                        df['TI'] = df['TI'].astype(str) ##
                        df['TO'] = df['TO'].astype(str) ##
                        df['VD'] = df['VD'].astype(str) ###
                        df['DD'] = df['DD'].astype(str) ####
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)        
                        # SPLITTING ART START DATE
                        A = df[df['AS'].str.contains('-')].copy()
                        a = df[~df['AS'].str.contains('-')].copy()
                        B = a[a['AS'].str.contains('/')].copy()
                        C = a[~a['AS'].str.contains('/')].copy()
                        E = C[C['AS'].str.contains(' ')].copy()
                        D = C[~C['AS'].str.contains(' ')].copy()
                        A[['Ayear', 'Amonth', 'Aday']] = A['AS'].str.split('-', expand = True)
                        B[['Ayear', 'Amonth', 'Aday']] = B['AS'].str.split('/', expand = True)
                        try:
                            D['AS'] = pd.to_numeric(D['AS'], errors='coerce')
                            D['AS'] = pd.to_datetime(D['AS'], origin='1899-12-30', unit='D', errors='coerce')
                            D['AS'] =  D['AS'].astype(str)
                            D[['Ayear', 'Amonth', 'Aday']] = D['AS'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['AS'] = pd.to_datetime(E['AS'],format='%d %m %Y', errors='coerce')
                            E['AS'] =  E['AS'].astype(str)
                            E[['Ayear', 'Amonth', 'Aday']] = E['AS'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E]) 
                        # SPLITTING DEATH DATE
                        A = df[df['DD'].str.contains('-')].copy()
                        a = df[~df['DD'].str.contains('-')].copy()
                        B = a[a['DD'].str.contains('/')].copy()
                        C = a[~a['DD'].str.contains('/')].copy()
                        E = C[C['DD'].str.contains(' ')].copy()
                        D = C[~C['DD'].str.contains(' ')].copy()
                        A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
                        B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)
                        try:
                            D['DD'] = pd.to_numeric(D['DD'], errors='coerce')
                            D['DD'] = pd.to_datetime(D['DD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['DD'] =  D['DD'].astype(str)
                            D[['Dyear', 'Dmonth', 'Dday']] = D['DD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['DD'] = pd.to_datetime(E['DD'],format='%d %m %Y', errors='coerce')
                            E['DD'] =  E['DD'].astype(str)
                            E[['Dyear', 'Dmonth', 'Dday']] = E['DD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])                     
                        # SORTING THE RETURN VISIT DATE
                        A = df[df['RD'].str.contains('-')].copy()
                        a = df[~df['RD'].str.contains('-')].copy()
                        B = a[a['RD'].str.contains('/')].copy()
                        C = a[~a['RD'].str.contains('/')].copy()
                        E = C[C['RD'].str.contains(' ')].copy()
                        D = C[~C['RD'].str.contains(' ')].copy()                     
                        #D = C[C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
                        #E = C[~C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()              
                        A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
                        B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
                        try:
                            D['RD'] = pd.to_numeric(D['RD'], errors='coerce')
                            D['RD'] = pd.to_datetime(D['RD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD'] =  D['RD'].astype(str)
                            D[['Ryear', 'Rmonth', 'Rday']] = D['RD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD'] = pd.to_datetime(E['RD'],format='%d %m %Y', errors='coerce')
                            E['RD'] =  E['RD'].astype(str)
                            E[['Ryear', 'Rmonth', 'Rday']] = E['RD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])                
                        #SORTING THE VD DATE
                        A = df[df['VD'].str.contains('-')].copy()
                        a = df[~df['VD'].str.contains('-')].copy()
                        B = a[a['VD'].str.contains('/')].copy()
                        C = a[~a['VD'].str.contains('/')].copy()
                        E = C[C['VD'].str.contains(' ')].copy()
                        D = C[~C['VD'].str.contains(' ')].copy()      
                        A[['Vyear', 'Vmonth', 'Vday']] = A['VD'].str.split('-', expand = True)
                        B[['Vyear', 'Vmonth', 'Vday']] = B['VD'].str.split('/', expand = True)
                        try:
                            D['VD'] = pd.to_numeric(D['VD'], errors='coerce')
                            D['VD'] = pd.to_datetime(D['VD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['VD'] =  D['VD'].astype(str)
                            D[['Vyear', 'Vmonth', 'Vday']] = D['VD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['VD'] = pd.to_datetime(E['VD'],format='%d %m %Y', errors='coerce')
                            E['VD'] =  E['VD'].astype(str)
                            E[['Vyear', 'Vmonth', 'Vday']] = E['VD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])  
                        df = df.copy()
                        #SORTING THE TO DATE
                        A = df[df['TO'].str.contains('-')].copy()
                        a = df[~df['TO'].str.contains('-')].copy()
                        B = a[a['TO'].str.contains('/')].copy()
                        C = a[~a['TO'].str.contains('/')].copy()
                        E = C[C['TO'].str.contains(' ')].copy()
                        D = C[~C['TO'].str.contains(' ')].copy()         
                        A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
                        B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
                        try:
                            D['TO'] = pd.to_numeric(D['TO'], errors='coerce')
                            D['TO'] = pd.to_datetime(D['TO'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TO'] =  D['TO'].astype(str)
                            D[['Tyear', 'Tmonth', 'Tday']] = D['TO'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TO'] = pd.to_datetime(E['TO'],format='%d %m %Y', errors='coerce')
                            E['TO'] =  E['TO'].astype(str)
                            E[['Tyear', 'Tmonth', 'Tday']] = E['TO'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])         
                    #SORTING THE TI DATE
                        A = df[df['TI'].str.contains('-')].copy()
                        a = df[~df['TI'].str.contains('-')].copy()
                        B = a[a['TI'].str.contains('/')].copy()
                        C = a[~a['TI'].str.contains('/')].copy()
                        E = C[C['TI'].str.contains(' ')].copy()
                        D = C[~C['TI'].str.contains(' ')].copy()         
                        A[['Tiyear', 'Timonth', 'Tiday']] = A['TI'].str.split('-', expand = True)
                        B[['Tiyear', 'Timonth', 'Tiday']] = B['TI'].str.split('/', expand = True)
                        try:
                            D['TI'] = pd.to_numeric(D['TI'], errors='coerce')
                            D['TI'] = pd.to_datetime(D['TI'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TI'] =  D['TI'].astype(str)
                            D[['Tiyear', 'Timonth', 'Tiday']] = D['TI'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TI'] = pd.to_datetime(E['TI'],format='%d %m %Y', errors='coerce')
                            E['TI'] =  E['TI'].astype(str)
                            E[['Tiyear', 'Timonth', 'Tiday']] = E['TI'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                        # SORTING THE RETURN VISIT DATE1
                        A = df[df['RD1'].str.contains('-')].copy()
                        a = df[~df['RD1'].str.contains('-')].copy()
                        B = a[a['RD1'].str.contains('/')].copy()
                        C = a[~a['RD1'].str.contains('/')].copy()
                        E = C[C['RD1'].str.contains(' ')].copy()
                        D = C[~C['RD1'].str.contains(' ')].copy()
                        A[['Royear', 'R1month', 'R1day']] = A['RD1'].str.split('-', expand = True)
                        B[['Royear', 'R1month', 'R1day']] = B['RD1'].str.split('/', expand = True)
                        try:
                            D['RD1'] = pd.to_numeric(D['RD1'], errors='coerce')
                            D['RD1'] = pd.to_datetime(D['RD1'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD1'] =  D['RD1'].astype(str)
                            D[['Royear', 'R1month', 'R1day']] = D['RD1'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD1'] = pd.to_datetime(E['RD1'],format='%d %m %Y', errors='coerce')
                            E['RD1'] =  E['RD1'].astype(str)
                            E[['Royear', 'R1month', 'R1day']] = E['RD1'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                        # SORTING THE RETURN VISIT DATE2
                        A = df[df['RD2'].str.contains('-')].copy()
                        a = df[~df['RD2'].str.contains('-')].copy()
                        B = a[a['RD2'].str.contains('/')].copy()
                        C = a[~a['RD2'].str.contains('/')].copy()
                        E = C[C['RD2'].str.contains(' ')].copy()
                        D = C[~C['RD2'].str.contains(' ')].copy()
                        A[['R2year', 'R2month', 'R2day']] = A['RD2'].str.split('-', expand = True)
                        B[['R2year', 'R2month', 'R2day']] = B['RD2'].str.split('/', expand = True)
                        try:
                            D['RD2'] = pd.to_numeric(D['RD2'], errors='coerce')
                            D['RD2'] = pd.to_datetime(D['RD2'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD2'] =  D['RD2'].astype(str)
                            D[['R2year', 'R2month', 'R2day']] = D['RD2'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD2'] = pd.to_datetime(E['RD2'],format='%d %m %Y', errors='coerce')
                            E['RD2'] =  E['RD2'].astype(str)
                            E[['R2year', 'R2month', 'R2day']] = E['RD2'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                        # SORTING THE LAST ENCOUNTER DATES
                        A = df[df['LD'].str.contains('-')].copy()
                        a = df[~df['LD'].str.contains('-')].copy()
                        B = a[a['LD'].str.contains('/')].copy()
                        C = a[~a['LD'].str.contains('/')].copy()
                        E = C[C['LD'].str.contains(' ')].copy()
                        D = C[~C['LD'].str.contains(' ')].copy()
                        A[['Lyear', 'Lmonth', 'Lday']] = A['LD'].str.split('-', expand = True)
                        B[['Lyear', 'Lmonth', 'Lday']] = B['LD'].str.split('/', expand = True)
                        try:
                            D['LD'] = pd.to_numeric(D['LD'], errors='coerce')
                            D['LD'] = pd.to_datetime(D['LD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['LD'] =  D['LD'].astype(str)
                            D[['Lyear', 'Lmonth', 'Lday']] = D['LD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['LD'] = pd.to_datetime(E['LD'],format='%d %m %Y', errors='coerce')
                            E['LD'] =  E['LD'].astype(str)
                            E[['Lyear', 'Lmonth', 'Lday']] = E['LD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                       
                        # SORTING THE FIRST ENCOUNTER
                        A = df[df['FE'].str.contains('-')].copy()
                        a = df[~df['FE'].str.contains('-')].copy()
                        B = a[a['FE'].str.contains('/')].copy()
                        C = a[~a['FE'].str.contains('/')].copy()
                        E = C[C['FE'].str.contains(' ')].copy()
                        D = C[~C['FE'].str.contains(' ')].copy()
                        A[['Fyear', 'Fmonth', 'Fday']] = A['FE'].str.split('-', expand = True)
                        B[['Fyear', 'Fmonth', 'Fday']] = B['FE'].str.split('/', expand = True)
                        try:
                            D['FE'] = pd.to_numeric(D['FE'], errors='coerce')
                            D['FE'] = pd.to_datetime(D['FE'], origin='1899-12-30', unit='D', errors='coerce')
                            D['FE'] =  D['FE'].astype(str)
                            D[['Fyear', 'Fmonth', 'Fday']] = D['FE'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['FE'] = pd.to_datetime(E['FE'],format='%d %m %Y', errors='coerce')
                            E['FE'] =  E['FE'].astype(str)
                            E[['Fyear', 'Fmonth', 'Fday']] = E['FE'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                        #BRINGING BACK THE / IN DATES
                        df['AS'] = df['AS'].astype(str)
                        df['RD'] = df['RD'].astype(str)
                        df['RD1'] = df['RD1'].astype(str)
                        df['RD2'] = df['RD2'].astype(str)
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['VD'] = df['VD'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        df['FE'] = df['FE'].astype(str)
            #             #Clearing NaT from te dates
                        df['AS'] = df['AS'].str.replace('NaT', '',regex=True)
                        df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
                        df['RD1'] = df['RD1'].str.replace('NaT', '',regex=True)
                        df['RD2'] = df['RD2'].str.replace('NaT', '',regex=True)
                        df['TI'] = df['TI'].str.replace('NaT', '',regex=True)
                        df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
                        df['VD'] = df['VD'].str.replace('NaT', '',regex=True)
                        df['DD'] = df['DD'].str.replace('NaT', '',regex=True)
                        df['LD'] = df['LD'].str.replace('NaT', '',regex=True)
                        df['FE'] = df['FE'].str.replace('NaT', '',regex=True)
                                    #SORTING THE VIRAL LOAD YEARS
                        df[['Vyear', 'Vmonth', 'Vday']] =df[['Vyear', 'Vmonth', 'Vday']].apply(pd.to_numeric, errors = 'coerce') 
                        df['Vyear'] = df['Vyear'].fillna(994)
                        a = df[df['Vyear']>31].copy()
                        b = df[df['Vyear']<32].copy()
                        #c = df[]
                        b = b.rename(columns={'Vyear': 'Vday2', 'Vday': 'Vyear'})
                        b = b.rename(columns={'Vday2': 'Vday'})
                        df = pd.concat([a,b])
                        dfa = df.shape[0]
                        #SORTING THE TI YEARS
                        df[['Tiyear', 'Tiday']] =df[['Tiyear','Tiday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Tiyear'] = df['Tiyear'].fillna(994)
                        a = df[df['Tiyear']>31].copy()
                        b = df[df['Tiyear']<32].copy()
                        b = b.rename(columns={'Tiyear': 'Tiday2', 'Tiday': 'Tiyear'})
                        b = b.rename(columns={'Tiday2': 'Tiday'})
                        df = pd.concat([a,b])
                        dfb = df.shape[0]
                        # #SORTING THE RETURN VISIT DATE YEARS
                        df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        df['Ryear'] = df['Ryear'].fillna(994)
                        a = df[df['Ryear']>31].copy()
                        b = df[df['Ryear']<32].copy()
                        b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
                        b = b.rename(columns={'Rday2': 'Rday'})
                        df = pd.concat([a,b])
                            #SORTING THE TRANSFER OUT DATE YEAR
                        df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
                        df['Tyear'] = df['Tyear'].fillna(994)
                        a = df[df['Tyear']>31].copy()
                        b = df[df['Tyear']<32].copy()
                        b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
                        b = b.rename(columns={'Tday2': 'Tday'})
                        df = pd.concat([a,b])         
                        #SORTING THE ART START YEARS
                        df[['Ayear', 'Amonth', 'Aday']] =df[['Ayear', 'Amonth', 'Aday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Ayear'] = df['Ayear'].fillna(994)
                        a = df[df['Ayear']>31].copy()
                        b = df[df['Ayear']<32].copy()
                        b = b.rename(columns={'Ayear': 'Aday2', 'Aday': 'Ayear'})
                        b = b.rename(columns={'Aday2': 'Aday'})
                        df = pd.concat([a,b])
                        dfe = df.shape[0]
                        #SORTING THE ART START YEARS
                        df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Dyear'] = df['Dyear'].fillna(994)
                        a = df[df['Dyear']>31].copy()
                        b = df[df['Dyear']<32].copy()
                        b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
                        b = b.rename(columns={'Dday2': 'Dday'})
                        df = pd.concat([a,b])
                        dfe = df.shape[0]
                        # #SORTING THE RETURN VISIT DATE1
                        df[['R1day', 'Royear']] = df[['R1day', 'Royear']].apply(pd.to_numeric, errors='coerce')
                        df['Royear'] = df['Royear'].fillna(994)
                        a = df[df['Royear']>31].copy()
                        b = df[df['Royear']<32].copy()
                        b = b.rename(columns={'Royear': 'R1day2', 'R1day': 'Royear'})
                        b = b.rename(columns={'R1day2': 'R1day'})
                        df = pd.concat([a,b])
                        # #SORTING THE RETURN VISIT DATE2
                        df[['R2day', 'R2year']] = df[['R2day', 'R2year']].apply(pd.to_numeric, errors='coerce')
                        df['R2year'] = df['R2year'].fillna(994)
                        a = df[df['R2year']>31].copy()
                        b = df[df['R2year']<32].copy()
                        b = b.rename(columns={'R2year': 'R2day2', 'R2day': 'R2year'})
                        b = b.rename(columns={'R2day2': 'R2day'})
                        df = pd.concat([a,b])
                        # #SORTING THE LAST ENCOUNTER
                        df[['Lday', 'Lyear']] = df[['Lday', 'Lyear']].apply(pd.to_numeric, errors='coerce')
                        df['Lyear'] = df['Lyear'].fillna(994)
                        a = df[df['Lyear']>31].copy()
                        b = df[df['Lyear']<32].copy()
                        b = b.rename(columns={'Lyear': 'Lday2', 'Lday': 'Lyear'})
                        b = b.rename(columns={'Lday2': 'Lday'})
                        df = pd.concat([a,b])
                        df = df.copy()
                        # #SORTING THE FIRST ENCOUNTER
                        df[['Fday', 'Fyear']] = df[['Fday', 'Fyear']].apply(pd.to_numeric, errors='coerce')
                        df['Fyear'] = df['Fyear'].fillna(994)
                        df = df.copy()
                        a = df[df['Fyear']>31].copy()
                        b = df[df['Fyear']<32].copy()
                        b = b.rename(columns={'Fyear': 'Fday2', 'Fday': 'Fyear'})
                        b = b.rename(columns={'Fday2': 'Fday'})
                        df = pd.concat([a,b])
                        df = df.copy()
                       
                        #CREATE WEEKS 
                        df['Rdaya'] = df['Rday'].astype(str).str.split('.').str[0]
                        df['Rmontha'] = df['Rmonth'].astype(str).str.split('.').str[0]
                        df['Ryeara'] = df['Ryear'].astype(str).str.split('.').str[0]
                        df['RETURN DATE'] = df['Rdaya'] + '/' + df['Rmontha'] + '/' + df['Ryeara']
                        df['RETURN DATE'] = pd.to_datetime(df['RETURN DATE'], format='%d/%m/%Y', errors='coerce')
                        #CREATING WEEEK FOR RETURN VISIT DATE
                        df['RWEEK'] = df['RETURN DATE'].dt.strftime('%V')
                        df['RWEEK'] = pd.to_numeric(df['RWEEK'], errors='coerce')
                        # df['RWEEK1'] = df['RWEEK'] + 13
                        df['RWEEK1'] = df['RWEEK'] - 39
                        #       #PARAMETERS TO SORT OUT FALSE TOs, USING LD AND T
                        df['Taya'] = df['Tday'].astype(str).str.split('.').str[0]
                        df['Tmontha'] = df['Tmonth'].astype(str).str.split('.').str[0]
                        df['Tyeara'] = df['Tyear'].astype(str).str.split('.').str[0]
                        df['TO DATE'] = df['Taya'] + '/' + df['Tmontha'] + '/' + df['Tyeara']
                        df['TO DATE'] = pd.to_datetime(df['TO DATE'], format='%d/%m/%Y', errors='coerce')
                       #LAST ENCOUTER TO DATES
                        df['Ldaya'] = df['Lday'].astype(str).str.split('.').str[0]
                        df['Lmontha'] = df['Lmonth'].astype(str).str.split('.').str[0]
                        df['Lyeara'] = df['Lyear'].astype(str).str.split('.').str[0]
                        df['LAST DATE'] = df['Ldaya'] + '/' + df['Lmontha'] + '/' + df['Lyeara']
                        df['LAST DATE'] = pd.to_datetime(df['LAST DATE'], format='%d/%m/%Y', errors='coerce')
                        df['DURA'] = round((df['LAST DATE']-df['TO DATE']).dt.days)
                        df['VR'] = pd.to_numeric(df['VR'], errors='coerce')
                        df['VR'] = df['VR'].fillna(20)

                        def oneager(a):
                             if a < 10:
                                  return '0-9'
                             elif a < 20:
                                  return '10-19'
                             elif a < 25:
                                  return '20-24'
                             else:
                                  return '25+'

                        def ager2(a):
                             if a <5:
                                  return '0-4'
                             elif a < 10:
                                  return '5-9'
                             elif a < 15:
                                  return '10-14'
                             elif a < 20:
                                  return '15-19'
                             elif a < 25:
                                  return '20-24'
                             elif a < 30:
                                  return '25-29'
                             elif a < 40:
                                  return '30-39'
                             elif a < 50:
                                  return '40-49'
                             elif a >49:
                                  return '50+'
                        df['AG'] = pd.to_numeric(df['AG'], errors='coerce')
                        df['AGE BAND'] = df['AG'].apply(ager2)
                        df['AGE_BANDS'] = df['AG'].apply(oneager)
                        daty = {
                             'AGE BAND': ['0-4', '5-9', '10-14', '15-19', '20-24',
                                          '25-29', '30-39', '40-49', '50+']
                        }
                        dfcomp = pd.DataFrame(daty)

                        oneyear = df.copy()

        ################################ PARAMETERS ############################################
                        #Q1 parameters
                        b1year = 2025 #one year ago the upper limit for bbtc, goes with lmonth and lday
                        b2year = 2024 #one year ago the lower limit for bbtc, goes with lmonth and lday

                        cyear = 2026  #curr year
                        cyp = 2027 # a year after
                        cyp1 = cyp +1
                        cmonth = 3 #last month of the qtr
                        cml = 2 # a month before
                        cmp = 4 # a month after
                        cday  = 4 #starting day
                        cdp = 5   # a day after
                        cdm = 3 #a day before
                        qmonths = [1,2,3] # months of the qtr

                        lmonth = 12 # last qtr month, used for txcur
                        lyear = 2025 # last qtr  year
                        lday = 4  #last qtr day
                        ldp = 5 # a day after
                        ldm = 3 # a day before

                        oyear = 2025  #the year for one year cohort

                        vyeara = 2025 # current vl year
                        vmonth = 3 # last vl month for those with (>) lowest limit, for has
                        vmm = 4  # a month after for those without (<) for lacks
                        vayear = 2025 #for art start date in vl, cutt of six months
                        vamonth = 10 # for art start date in vl cutt off sixmonth <
                        qt6 = 'Q4' #the qtr for six months ago

                        fmonth =  1 # first month of this current qtr

                        qtr = 'Q1' #the previous qtr column from clusters.csv

                        #NEXT QTR PARAMETERS
                        nyear = 2026
                        nmonth = 4 #startin month of next qtr

                        #6months ago months
                        q6months = [7,8,9]
                        
                        q6year = 2025
                        #####
                        year24 = 2024 # FOR 24 months cohort

                        #MONTHS FOR COLUMN 1 IN COHORTS
                        mths6 = 'JUL-SEP'
                        onyr = 'JAN-MAR'
                        
                        
        
##################################################################################################################################
                        #POTENTIAL TXCUR ALTER... 
                        df[['Rmonth', 'Rday', 'Ryear']] = df[['Rmonth', 'Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        df25 = df[df['Ryear']>lyear].copy()
                        df24 = df[df['Ryear'] == lyear].copy()
                        df24[['Rmonth', 'Rday']] = df24[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                        df24 = df24[((df24['Rmonth']>lmonth) | ((df24['Rmonth']==lmonth) & (df24['Rday']>ldm)))].copy()
                        df = pd.concat([df25, df24]).copy()
                        
                        df = df.copy()
                        # DSD = df.copy()
        
                        #REMOVE TO of the last reporting month
                        df[ 'Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
                        dfto = df[df['Tyear']!=994].copy() #HAVE TOs
                        dfnot = df[df['Tyear'] == 994].copy() #NO TO
        
                       #REMOVE THE TO
                        dfto[['Ryear', 'Rmonth']] = dfto[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                        dftoy = dfto[((dfto['Ryear']!=lyear) |((dfto['Ryear']==lyear) & (dfto['Rmonth']>lmonth)))].copy() #OTHERS WOULD BE FALSE TOs, even those made last Q since they were brought as false if their RRDs were this year
                        
                        dftox = dfto[((dfto['Ryear']==lyear) & (dfto['Rmonth']==lmonth))].copy() #CLIENTS WITH RD OF REPORTING MONTH, DIDN'T RETURN BUT WERE TO LATER
                        dftox[['Tyear', 'Tmonth']] = dftox[['Tyear', 'Tmonth']].apply(pd.to_numeric, errors='coerce')
                        dftox = dftox[((dftox['Tyear']==lyear) & (dftox['Tmonth']>lmonth))].copy()

        
                        df = pd.concat([dftoy,dfnot])
                        
                        df = df.copy()
                        if dftox.shape[0]>0:
                            df = pd.concat([df,dftox])
                        else:
                            df =df.copy()
                        #REMOVE the dead of the reporting month
                        df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
                        dfdd = df[df['Dyear']!=994].copy()
                        dfnot = df[df['Dyear'] == 994].copy()
                        #THOSE WHO DIED BEFORE FIRST MONTH OF THE Q
                        dfdd[['Dyear', 'Dmonth']] = dfdd[['Dyear', 'Dmonth']].apply(pd.to_numeric, errors='coerce')
                        dfdd = dfdd[((dfdd['Dyear']>lyear) |((dfdd['Dyear']==lyear) & (dfdd['Dmonth']>lmonth)))].copy() #DOESN'T MAKE SENSE
                        
                        df = pd.concat([ dfdd,dfnot])

                        #REMOVE THE TI AND NEWS OF NEXT QTR, USEFUL DURING DATA COLLECTION
                        df[['Ayear', 'Amonth']] = df[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        df = df[((df['Ayear']<nyear) |((df['Ayear']==nyear) & (df['Amonth']<nmonth)))].copy() #ART START AFTER FIRST MONTH OF THE QTR, DOESN'T MAKE SENSE FOR THEM TO BE IN POT CURR
                        df[['Tiyear', 'Timonth']] = df[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
                        df = df[((df['Tiyear']<nyear) |((df['Tiyear']==nyear) & (df['Timonth']<nmonth)))].copy() #TI AFTER FIRST MONTH OF THE QTR, DOESN'T MAKE SENSE FOR THEM TO BE IN POT CURR
                        # st.write(df.shape[0])
                        #dfn = df[((df['Tiyear']>nyear) |((df['Tiyear']==nyear) & (df['Timonth']==nmonth)))].copy() #TI AFTER FIRST MONTH OF THE QTR, DOESN'T MAKE SENSE FOR THEM TO BE IN POT CURR
                        
                        
                        df['AG'] = pd.to_numeric(df['AG'],errors='coerce')
                        
                        dar = {
                                'AGE BAND': ['NO DATA' ,'','KINDLY CONFIRM'],
                                'F': ['NO DATA','',''],
                                'M': ['NO DATA', '','']
                                    } 
                        dfdar = pd.DataFrame(dar)
                        df['VR']  = pd.to_numeric(df['VR'], errors='coerce')
                        df['VR'] = df['VR'].fillna(30)
                        def mmd(a):
                             if a < 90:
                                  return '<3'
                             elif a < 179:
                                  return '3-5'
                             else:
                                  return '6 mths'  
                        df['ARVD'] = pd.to_numeric(df['ARVD'], errors='coerce')  
                        df['AG'] = pd.to_numeric(df['AG'], errors='coerce') 
                        df['AGE BAND2'] = df['AG'].apply(ager2)          
                        df = df.copy()
                        dfrt = df.copy()
                        df6m = df.copy()

                        newly = df.copy()
                        
                        tri = df.copy()
                        
                        newly[['Ayear', 'Amonth']] = newly[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
    
                        newly = newly[((newly['Ayear']==cyear) & (newly['Amonth'].isin(qmonths)))].copy()
                        newly['Tiyear'] = pd.to_numeric(newly['Tiyear'], errors='coerce')
                        newly = newly[newly['Tiyear']==994].copy()

                        
                        dfhc01 = newly[['ART', 'GD', 'AGE BAND', 'CD' ]].copy()
                        base = newly.copy()

                        tri[['Tiyear', 'Timonth']] = tri[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
                        tri = tri[((tri['Tiyear']==cyear) & (tri['Timonth'].isin(qmonths)))].copy()
                        dfhc07 = tri[['ART', 'GD', 'AGE BAND' ]].copy()
                        
                        base['CD'] = pd.to_numeric(base['CD'], errors='coerce')
                        hbase = base[base['CD'].notna()].copy()
                        
                        hbase['CD'] = pd.to_numeric(hbase['CD'], errors='coerce')
                        lbase = hbase[hbase['CD']<200].copy()
                        dfhc08 = hbase[['ART', 'GD', 'AGE BAND' ]].copy()
                        dfhc09 = lbase[['ART', 'GD', 'AGE BAND' ]].copy()
                    
                        
                        df6m[['Ayear', 'Amonth']] = df6m[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        df6m = df6m[((df6m['Ayear']==q6year) & (df6m['Amonth'].isin(q6months)))].copy()
                        dfhc26_Elig = df6m[['ART', 'GD', 'AGE BAND','A' ]].copy()


                        df6m[['Vyear', 'Vmonth']] = df6m[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        df6m = df6m[((df6m['Vyear']==cyear) & (df6m['Vmonth'].isin(qmonths)))].copy()

                        dfhc26_test = df6m[['ART', 'GD', 'AGE BAND','A' ]].copy()

                        df6m['VR']  = pd.to_numeric(df6m['VR'], errors='coerce')
                        df6m = df6m[df6m['VR']<1000].copy()

                        dfhc26_sup = df6m[['ART', 'GD', 'AGE BAND','A' ]].copy()


                       
        ##A             
                        #QUARTERLY TX ML\
                        dfcurr = df.copy()
                        # #DEAD
                        dfcurr['Dyear'] = pd.to_numeric(dfcurr['Dyear'], errors='coerce')
                        deadq = dfcurr[dfcurr['Dyear']!=994].copy()  #THE DEAD
                        dfhc14_DIED = deadq[['ART','GD', 'AGE BAND']].copy()
                        dfcurr = dfcurr[dfcurr['Dyear']==994].copy() #REMOVED THE DEAD
    
                        # ####TO
                        dfcurr['Tyear'] = pd.to_numeric(dfcurr['Tyear'], errors='coerce')
                        dfcurra = dfcurr[dfcurr['Tyear']==994].copy()  #NO TO 

                        dfcto = dfcurr[dfcurr['Tyear']!=994].copy() #HAS TOs, false TOs that are txml and active
                        
                        dfcto[['Ryear', 'Rmonth', 'RWEEK']] =  dfcto[['Ryear', 'Rmonth', 'RWEEK']].apply(pd.to_numeric) #TO USE WEEKS FOR NOW    
                        dfctoF = dfcto[ ((dfcto['Ryear']> cyear) | ((dfcto['Ryear'] ==cyear) & (dfcto['Rmonth']>cmonth))) ].copy()
                        #dfctoF = dfcto[ ((dfcto['Ryear']> cyear) | ((dfcto['Ryear'] ==cyear) & (dfcto['RWEEK']>=thisweek))) ].copy()
        
                        dfctoT = dfcto[ ((dfcto['Ryear']< cyear) | ((dfcto['Ryear'] ==cyear) & (dfcto['Rmonth']<cmp))) ].copy() #HAVE TRUE TOS, OLD TOS
                        #dfctoT = dfcto[ ((dfcto['Ryear']< cyear) | ((dfcto['Ryear'] ==cyear) & (dfcto['RWEEK']<thisweek))) ].copy() #HAVE TRUE TOS, OLD TOS
                        dfctoT[['Tyear', 'Tmonth']] = dfctoT[['Tyear', 'Tmonth']].apply(pd.to_numeric, errors='coerce')
                        
                        #OLD TO VS TO OF THE Q        
                        dfctold = dfctoT[((dfctoT['Tyear']<cyear)| ((dfctoT['Tyear'] ==cyear) & (dfctoT['Tmonth']<fmonth)))].copy() #OLD TOs, MADE BEFORE FIRST MONTH OF THE QTR
                        
                        dfctoT = dfctoT[((dfctoT['Tyear'] ==cyear) & (dfctoT['Tmonth'].isin(qmonths)))].copy() #TOs made this Q, ARE THE TOs
                        
                        dfctold[['Rmonth', 'Rday']] = dfctold[['Rmonth', 'Rday']].apply(pd.to_numeric)

                        #OLD TOs returned, that are still active
                        dfctox = dfctold[((dfctold['Rmonth'] ==cmonth) & (dfctold['Rday']>cdm))].copy() #ADD THEM BACK TO CURR
                        #dfctox = dfctold[dfctold['RWEEK']>=thisweek].copy() #ADD THEM BACK TO CURR
        
                        #OLD TOs THAT ARE LOST
                        dfctoyx = dfctold[((dfctold['Rmonth']< cmonth) | ((dfctold['Rmonth']==cmonth) & (dfctold['Rday']<cday)))].copy() #ADD THEM TO LOST
                        #dfctoyx = dfctold[dfctold['RWEEK']<thisweek].copy() #ADD THEM TO LOST
                        dfctoyx['DURA'] = pd.to_numeric(dfctoyx['DURA'], errors = 'coerce')
                        dfctoy = dfctoyx[dfctoyx['DURA']>0].copy() #RETURNED GOT LOST, ADD TO TXML
                        dfctoyz= dfctoyx[dfctoyx['DURA']<1].copy() #NEVER RETURNED, IS A TO
        
                        #stos vs not
                        dfctoT['NTO'] = dfctoT['NTO'].astype(str)
                        words = ['sto', 's\\.t\\.o', 'st\\.o', 'self']
                        pattern = '|'.join(words)
                        dfsto = dfctoT[dfctoT['NTO'].str.contains(pattern, case=False, na=False)]
                        dfsto = dfsto.copy()
                        dfsto['A'] = pd.to_numeric(dfsto['A'], errors= 'coerce')
                        dfctoT['A'] = pd.to_numeric(dfctoT['A'], errors= 'coerce')
                        dfTO =  dfctoT[~dfctoT['A']. isin(dfsto['A'])].copy()
                        dfTO = pd.concat([dfTO, dfctoyz])
                        dfTO = dfTO.copy()
                        dfhc14_TO = dfsto[['ART', 'AGE BAND', 'GD']].copy()
                      
                        dfcur = pd.concat([dfcurra, dfctoF])
                        dfcur = dfcur.copy()

                        #ON APPT
                        dfcur[['Rday','Rmonth', 'Ryear']] = dfcur[['Rday','Rmonth', 'Ryear']].apply(pd.to_numeric, errors = 'coerce')
                         
                        lacks = dfcur[((dfcur['Vyear']< vyeara) | ((dfcur['Vyear'] ==vyeara) & (dfcur['Vmonth']<vmm)))].copy()
                        lacks[['Ayear', 'Amonth']] = lacks[['Ayear', 'Amonth']].apply(pd.to_numeric, errors ='coerce')
                        lacks = lacks[((lacks['Ayear']<vayear) |((lacks['Ayear']==vayear)& (lacks['Amonth'] <vamonth)))].copy()
                        lacks = lacks[lacks['Ayear']!=994].copy()
                        # lacks = lacks[lacks['Tyear']==994].copy()
         
                        
                        dfcur[['Ryear', 'Rmonth']] = dfcur[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors ='coerce')
                        #LOST LAST QTR
                        currlosta = dfcur[((dfcur['Ryear'] == lyear) & (dfcur['Rmonth']==lmonth))].copy()
            
                        #LOST THIS QTR
                        curlostb = dfcur[((dfcur['Ryear'] == cyear) & (dfcur['Rmonth'].isin(qmonths)))].copy() #LOST THIS QTR
                        curlostb[['Ryear', 'Rmonth', 'Rday', 'RWEEK']] = curlostb[['Ryear', 'Rmonth', 'Rday','RWEEK']].apply(pd.to_numeric, errors ='coerce')
                        curlostc = curlostb[ ((curlostb['Rmonth']<cmonth) |(( curlostb['Rmonth']== cmonth) & (curlostb['Rday']<cday)))].copy()
                        #curlostc = curlostb[curlostb['RWEEK']<thisweek].copy()
        
                        #currlost = pd.concat([curlostc, dfctoy]) #x will be the TOs that returned and got lost, in second q include curlosta
                        framers = [curlostc, dfctoy, currlosta]
                        framers = [f for f in framers if not f.empty]
                        if framers:  # at least one is non-empty
                            currlost = pd.concat(framers, ignore_index=True)
                        else:       # all three are empty
                            currlost = pd.DataFrame()
                            
                        currlost = currlost.copy()
                        if currlost.shape[0] ==0:
                             dfhc14_LOST = pd.DataFrame(columns=['ART', 'GD', 'AGE BAND'])
                        else:
                            dfhc14_LOST = currlost[['ART', 'GD', 'AGE BAND']].copy()
                        
                        cur26 = dfcur[dfcur['Ryear'] >cyear].copy() #ACTIVE NEXT OTHER YEARS
                        cur25 = dfcur[dfcur['Ryear'] == cyear].copy() # ACTIVE THIS YEAR
                        cur25[['Ryear', 'Rmonth', 'Rday']] = cur25[['Ryear', 'Rmonth', 'Rday']].apply(pd.to_numeric, errors ='coerce')
                        cur25 = cur25[ ((cur25['Rmonth']>cmonth) |(( cur25['Rmonth']==cmonth) & (cur25['Rday']>cdm)))].copy()
       
                        frames = [cur25, cur26, dfctox]
                        frames = [f for f in frames if not f.empty]
                        if frames:  # at least one is non-empty
                            dfcur = pd.concat(frames, ignore_index=True)
                        else:       # all three are empty
                            dfcur = pd.DataFrame()
                        if dfcur.shape[0] ==0:
                            st.warning('NO ACTIVE CLIENTS IN THIS EXTRACT, TRY MANUAL FILTERING WITH YOUR RD COLUMN TO VERIFY FOR YOUR SELF')
                            st.stop()
                        else:
                            pass
                        dfcur[['Ayear', 'Amonth']] = dfcur[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        dfcur = dfcur[((dfcur['Ayear'] <cyear) | ((dfcur['Ayear'] == cyear) & dfcur['Amonth']<cmp))].copy() #REMOVES TX NEW DATA COLL MONTH
                        dfcur['Ryear'] =  pd.to_numeric(dfcur['Ryear'], errors='coerce')
                        dfcur = dfcur[dfcur['Ryear'] <cyp1].copy() #REMOVES EXTREME YEARS
        
                        # dfcur[['Tiyear', 'Timonth']] = dfcur[['Tiyear', 'Timonth']].apply(pd.to_numeric, errors='coerce')
                        # dfcur = dfcur[((dfcur['Tiyear'] <cyear) | ((dfcur['Tiyear'] == cyear) & dfcur['Timonth']<cmp))].copy() #REMOVES TX NEW DATA COLL MONTH
                        dfcur = dfcur.copy()
                        dfacv = dfcur.copy()
    
                        a1 = dfcur.shape[0]
                        # st.write(dfcur)
                        dfcrt = dfcur.copy()  #FIND CURR
                        dfctct = dfcur.copy()
                        dfcrt['ARVL'] = dfcrt['ARVL'].astype(str)
                        dfcrt['ARVL'] = dfcrt['ARVL'].str.replace('NaT', 'FIRST LINE ARV REGIMEN')
                        dfcrt['ARVL'] = dfcrt['ARVL'].str.replace('nan', 'FIRST LINE ARV REGIMEN')
                        dfcrt['ARVL'] = dfcrt['ARVL'].fillna('FIRST LINE ARV REGIMEN')
                        
                        dfst = dfcrt[dfcrt['ARVL']=='FIRST LINE ARV REGIMEN'].copy()
                        dfscd = dfcrt[dfcrt['ARVL']=='SECOND LINE ARV REGIMEN'].copy()
                        dfhc11a = dfst[['ART', 'GD', 'AGE BAND' ]].copy()
                        dfhc11b = dfscd[['ART', 'GD', 'AGE BAND' ]].copy()
                        dfctct[['Lyear', 'Lmonth']] = dfctct[['Lyear', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                        dfct = dfctct[((dfctct['Lyear']==cyear) & (dfctct['Lmonth'].isin(qmonths)))].copy()
                        
                        dfhc11d = dfct[['ART', 'GD', 'AGE BAND' ]].copy()  

                        dfacv[['Lyear', 'Lmonth']] = dfacv[['Lyear', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                        scr = dfacv[((dfacv['Lyear']==cyear) & (dfacv['Lmonth'].isin(qmonths)))].copy()
                        dfhc18_scr = scr[['ART', 'GD', 'AGE BAND','A' ]].copy()
                        

                        dfhc18_scr['A'] = pd.to_numeric(dfhc18_scr['A'], errors='coerce')
                        newly['A'] = pd.to_numeric(newly['A'], errors='coerce')
                        dfhc19_scr = dfhc18_scr[~dfhc18_scr['A'].isin(newly['A'] )].copy()
                                     
                        
                        b1 = dfst.shape[0]
                        c1 = dfscd.shape[0]
                        d1 = a1-b1-c1

                        st.write(f'**TX CUR IS {a1}**')

                        if d1 ==0:
                             pass
                        else:
                             st.warning('FIRST LINE + SECOND LINE NOT EQUAL TO CURR')
                             st.stop()


                        
                        dfcur['ARVD'] = pd.to_numeric(dfcur['ARVD'], errors = 'coerce')
                        dfcur['MMD'] = dfcur['ARVD'].apply(mmd)
                        df3 = dfcur[dfcur['MMD']=='<3'].copy()
                        df5 = dfcur[dfcur['MMD']=='3-5'].copy()
                        df6 = dfcur[dfcur['MMD']=='6 mths'].copy()
                        e1 = df3.shape[0]
                        e2 = df5.shape[0]
                        e3 = df6.shape[0]
                        dfhc12_3 = df3[['ART', 'GD', 'AGE BAND' ]].copy()
                        dfhc12_5 = df5[['ART', 'GD', 'AGE BAND' ]].copy()
                        dfhc12_6 = df6[['ART', 'GD', 'AGE BAND' ]].copy()
 
                        e4 = a1-e1-e2-e3
                        if e4 ==0:
                             pass
                        else:
                             st.warning('MMDs not equal')
                             st.stop()
        
                        # #VL SECTION 
                        dfcur[['Ayear', 'Amonth']] = dfcur[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        dfcur['DSD'] =  dfcur['DSD'].astype(str)
                        dmap = {
                             'Community Drug Distribution Point': 'GMH',
                             'Community Client Led ART Delivery': 'GMC',
                             'Facility Based Groups':'GMH',
                             'Fast Track Drug Refill':'IMF',
                             'Facility Based Individual Management': 'IMF'

                        }
                        dfcur['DSD'] = dfcur['DSD'].astype(str)
                        dfcur['DSD'] = dfcur['DSD'].fillna('Fast Track Drug Refill')
                        dfcur['DSD'] = dfcur['DSD'].str.replace('NaT','Fast Track Drug Refill')
                        dfcur['DSD'] = dfcur['DSD'].str.replace('nan','Fast Track Drug Refill')
                        dfcur['DSDM'] = dfcur['DSD'].map(dmap)
                        dfhc43 = dfcur[['ART', 'AGE BAND', 'GD','DSDM']].copy()    
        
                        # #VL SECTION 
                        dfcur[['Ayear', 'Amonth']] = dfcur[['Ayear', 'Amonth']].apply(pd.to_numeric, errors='coerce')
                        has = dfcur[((dfcur['Ayear'] < vayear) | ((dfcur['Ayear'] == vayear) & (dfcur['Amonth']<vamonth)))].copy()
                        dfhc29_elig = has[['ART', 'AGE BAND', 'GD']].copy()
                        #BLED
                        has[['Vyear', 'Vmonth']] = has[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        has = has[((has['Vyear'] > vyeara) | ((has['Vyear'] == vyeara) & (has['Vmonth']> vmonth)))].copy()
                        dfhc29_test = has[['ART', 'AGE BAND', 'GD']].copy()
                        has['VR'] = pd.to_numeric(has['VR'], errors = 'coerce')
                        has = has[has['VR']<1000].copy()
                        dfhc29_sup = has[['ART', 'AGE BAND', 'GD']].copy()                      
        
                        #ONE YEAR COHORT
                        sixmths = oneyear.copy()
                        t24mths = oneyear.copy()
                        datx = {'AGE_BANDS': ['0-9', '10-19', '20-24', '25+'],
                                '1': [0,0,0,0]    
                        }
                        datx = pd.DataFrame(datx)

                        daty = {'AGE_BANDS': ['0-9', '10-19', '20-24', '25+'],
                                '1': ['10-12','10-12','10-12','10-12']    
                        }
                        datx = pd.DataFrame(datx)
                        daty = pd.DataFrame(daty)
###################################################################6 MTHS
                        if sixmths.shape[0]>1:
                            sixmths[['Ayear', 'Amonth']] = sixmths[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                    
                            new6 = sixmths[((sixmths['Ayear']== q6year) & (sixmths['Amonth'].isin(q6months)))].copy()
                        else:
                             df2a = dfdar.copy() 
                             
                        if new6.shape[0]>1:   
                            new6[['Tiyear']] = new6[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                            
                            
                            orig6 = new6[new6['Tiyear']==994].copy()
                            st.write(orig6)
                            if orig6.shape[0]>0:
                                orig6['CD'] = pd.to_numeric(orig6['CD'], errors='coerce')
                                #  orig6a = orig6[['AGE_BANDS', 'ART']].copy()
                                orig6a = orig6.groupby('AGE_BANDS')['ART'].size().reset_index()
                                orig6a = orig6a.rename(columns={'ART':'2'})
                                orig6l = orig6[orig6['CD']<200].copy()
                                if orig6l.shape[0]>0:
                                        orig6l = orig6l.groupby('AGE_BANDS')['ART'].size().reset_index()
                                        orig6l = orig6l.rename(columns={'ART':'3'})
                                else:
                                    orig6l = datx.rename(columns={'1':'3'})         
                            else:
                                 orig6a = datx.rename(columns={'1':'2'})
                                 orig6l = datx.rename(columns={'1':'3'})
                            
                            tin = new6[new6['Tiyear']!=994].copy()
                            if tin.shape[0]>0:
                                 tina6 = tin.groupby('AGE_BANDS')['ART'].size().reset_index()
                                 tina6 = tina6.rename(columns={'ART': '4'})
                            else:
                                 tina6 = datx.rename(columns ={'1':'4'})
                            new6['Dyear'] = pd.to_numeric(new6['Dyear'], errors='coerce')
                            new6dead = new6[new6['Dyear']!=994].copy()
                            
                            new6 = new6[new6['Dyear']==994].copy() #AFTER REMOVING THE DEAD
                            #activ6E 1 YEAR
                            new6['A'] = pd.to_numeric(new6['A'], errors='coerce')
                            dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                            new6cur = new6[new6['A'].isin(dfcur['A'])].copy()
                            
                            new6lost = new6[~new6['A'].isin(dfcur['A'])].copy()
                            new6lost['Tyear'] = pd.to_numeric(new6lost['Tyear'], errors='coerce')
                            new6losto = new6lost[new6lost['Tyear']!=994].copy()
                            st.write('HERE')
                            
            
                            new6lost = new6lost[new6lost['Tyear']==994].copy()
                            
                            
                            if new6losto.shape[0] > 0:
                                    out6 = new6losto.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    out6 = out6.rename(columns={'ART':'5'})
                                    dato = daty['AGE_BANDS']
                                    out6 = pd.merge(dato, out6, on ='AGE_BANDS', how='left')
                                    out6['5'] = out6['5'].fillna(0)
                                    st.write(out6)
                                    
                            else:
                                    out6 = datx.rename(columns={'1':'5'})
                            
                            stop = datx.rename(columns = {'1':'7'})

                            if new6dead.shape[0] > 0:
                                    dead6 = new6dead.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    dead6 = dead6.rename(columns={'ART':'8'})           
                            else:
                                    dead6 = datx.rename(columns={'1':'8'})

                            miss = datx.rename(columns = {'1':'9'})

                            if new6lost.shape[0] > 0:
                                    lost6 = new6lost.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    lost6 = lost6.rename(columns={'ART':'10'})
                                    
                            else:
                                    lost6 = datx.rename(columns={'1':'10'})

                            if new6cur.shape[0] > 0:
                                    activ6 = new6cur.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    new6cur['VR'] = pd.to_numeric(new6cur['VR'], errors='coerce')
                                    sup6 = new6cur[(new6cur['VR'] < 1000) & (new6cur['Vday'].notnull())].copy()
                                    
                                    activ6 = activ6.rename(columns={'ART':'11'})
                                    if sup6.shape[0] > 0:
                                         sup6 = sup6.groupby('AGE_BANDS')['ART'].size().reset_index()
                                         sup6 = sup6.rename(columns={'ART':'12'})
                                    else:
                                         sup6 = datx.rename(columns={'1':'12'})
                                    
                            else:
                                    activ6 = datx.rename(columns={'1':'11'})
                            df2a = pd.merge(daty,orig6a, on='AGE_BANDS', how='left')
                            df2a['2'] = df2a['2'].fillna(0)

                            df2a = pd.merge(df2a,orig6l, on='AGE_BANDS', how='left')
                            df2a['3'] = df2a['3'].fillna(0)

                            df2a = pd.merge(df2a,tina6, on='AGE_BANDS', how='left')
                            df2a['4'] = df2a['4'].fillna(0)

                            df2a = pd.merge(df2a,out6, on='AGE_BANDS', how='left')
                            try:
                                df2a['5'] = df2a['5'].fillna(0)
                            except:
                                df2a['5'] = 0

                            df2a = pd.merge(df2a,stop, on='AGE_BANDS', how='left')
                            df2a['7'] = df2a['7'].fillna(0)

                            df2a = pd.merge(df2a,dead6, on='AGE_BANDS', how='left')
                            df2a['8'] = df2a['8'].fillna(0)

                            df2a = pd.merge(df2a,miss, on='AGE_BANDS', how='left')
                            df2a['9'] = df2a['9'].fillna(0)

                            df2a = pd.merge(df2a,lost6, on='AGE_BANDS', how='left')
                            df2a['10'] = df2a['10'].fillna(0)
                            try:
                                 ay = activ6.shape[0]
                            except:
                                 activ6 = pd.DataFrame(columns=['AGE_BANDS'])                          
 
                            df2a = pd.merge(df2a,activ6, on='AGE_BANDS', how='left')
                            
                            df2a['11'] = df2a['11'].fillna(0)
                            try:
                                 av = sup6.shape[0]
                            except:
                                 sup6 = pd.DataFrame(columns=['AGE_BANDS', '12'])

                            df2a = pd.merge(df2a,sup6, on='AGE_BANDS', how='left')
                            df2a['12'] = df2a['12'].fillna(0)

                            df2a['6'] = df2a['2'] + df2a['4'] - df2a['5']

                            df2a['13'] = np.where(df2a['6'] == 0,0,np.floor(((df2a['11'] / df2a['6']) * 100) + 0.5)).astype(int)
                            df2a['1'] = df2a['1'].astype(str)
                            df2a['1'] = df2a['1'].str.replace('10-12', mths6)
                            df2a = df2a[['AGE_BANDS', '1', '2', '3', '4', '5','6', '7', '8', '9','10','11', '12', '13']].copy()
                            df2a = df2a.set_index('AGE_BANDS')                                            
                                      
                        else:
                             df2a = dfdar.copy()
#################################################################ONE YEAR
                        if oneyear.shape[0]>1:
                            oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                            
                            new = oneyear[((oneyear['Ayear']== oyear) & (oneyear['Amonth'].isin(qmonths)))].copy()
                        else:
                             df1 = dfdar.copy()   

                        if new.shape[0]>1: 
                            new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                            
                            orig = new[new['Tiyear']==994].copy()
                            if orig.shape[0]>0:
                                orig['CD'] = pd.to_numeric(orig['CD'], errors='coerce')
                                #  origa = orig[['AGE_BANDS', 'ART']].copy()
                                origa = orig.groupby('AGE_BANDS')['ART'].size().reset_index()
                                origa = origa.rename(columns={'ART':'2'})
                                origl = orig[orig['CD']<200].copy()
                                if origl.shape[0]>0:
                                        origl = origl.groupby('AGE_BANDS')['ART'].size().reset_index()
                                        origl = origl.rename(columns={'ART':'3'})
                                else:
                                    origl = datx.rename(columns={'1':'3'})         
                            else:
                                 origa = datx.rename(columns={'1':'2'})
                                 origl = datx.rename(columns={'1':'3'})
                            
                            tin = new[new['Tiyear']!=994].copy()
                            if tin.shape[0]>0:
                                 tina = tin.groupby('AGE_BANDS')['ART'].size().reset_index()
                                 tina = tina.rename(columns={'ART': '4'})
                            else:
                                 tina = datx.rename(columns ={'1':'4'})
                            #one =new.shape[0]                        
                            #LOSSES
                                #LOSSES
                            new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                            newdead = new[new['Dyear']!=994].copy()
                            
                            new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD
                            #ACTIVE 1 YEAR
                            new['A'] = pd.to_numeric(new['A'], errors='coerce')
                            dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                            newcur = new[new['A'].isin(dfcur['A'])].copy()
                            
                            newlost = new[~new['A'].isin(dfcur['A'])].copy()
                            newlost['Tyear'] = pd.to_numeric(newlost['Tyear'], errors='coerce')
                            newlosto = newlost[newlost['Tyear']!=994].copy()
            
                            newlost = newlost[newlost['Tyear']==994].copy()
                            
                            if newlosto.shape[0] > 0:
                                    out = newlosto.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    out = out.rename(columns={'ART':'5'})
                                    dato = daty['AGE_BANDS']
                                    out = pd.merge(dato, out, on ='AGE_BANDS', how='left')
                                    out['5'] = out['5'].fillna(0)
                            else:
                                    out = datx.rename(columns={'1':'5'})
                            
                            stop = datx.rename(columns = {'1':'7'})

                            if newdead.shape[0] > 0:
                                    deadz = newdead.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    deadz = deadz.rename(columns={'ART':'8'})           
                            else:
                                    deadz = datx.rename(columns={'1':'8'})

                            miss = datx.rename(columns = {'1':'9'})

                            if newlost.shape[0] > 0:
                                    lostx = newlost.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    lostx = lostx.rename(columns={'ART':'10'})
                                    
                            else:
                                    lostx = datx.rename(columns={'1':'10'})

                            if newcur.shape[0] > 0:
                                    activ = newcur.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    newcur['VR'] = pd.to_numeric(newcur['VR'], errors='coerce')
                                    supx = newcur[newcur['VR'] <1000].copy()
                                    activ = activ.rename(columns={'ART':'11'})
                                    if supx.shape[0] > 0:
                                         supx = supx.groupby('AGE_BANDS')['ART'].size().reset_index()
                                         supx = supx.rename(columns={'ART':'12'})
                                    else:
                                         supx = datx.rename(columns={'1':'12'})
                                    
                            else:
                                    newcur = datx.rename(columns={'1':'11'})
                            df1 = pd.merge(daty,origa, on='AGE_BANDS', how='left')
                            df1['2'] = df1['2'].fillna(0)

                            df1 = pd.merge(df1,origl, on='AGE_BANDS', how='left')
                            df1['3'] = df1['3'].fillna(0)

                            df1 = pd.merge(df1,tina, on='AGE_BANDS', how='left')
                            df1['4'] = df1['4'].fillna(0)


                            df1 = pd.merge(df1,out, on='AGE_BANDS', how='left')
                            try:
                                df2a['5'] = df2a['5'].fillna(0)
                            except:
                                df2a['5'] = 0

                            df1 = pd.merge(df1,stop, on='AGE_BANDS', how='left')
                            df1['7'] = df1['7'].fillna(0)

                            df1 = pd.merge(df1,deadz, on='AGE_BANDS', how='left')
                            df1['8'] = df1['8'].fillna(0)

                            df1 = pd.merge(df1,miss, on='AGE_BANDS', how='left')
                            df1['9'] = df1['9'].fillna(0)

                            df1 = pd.merge(df1,lostx, on='AGE_BANDS', how='left')
                            df1['10'] = df1['10'].fillna(0)
                            try:
                                 at = activ.shape[0]
                            except:
                                 activ = pd.DataFrame(columns=['AGE_BANDS'])   

                            df1 = pd.merge(df1,activ, on='AGE_BANDS', how='left')
                            df1['11'] = df1['11'].fillna(0)
                            try:
                                 avt = supx.shape[0]
                            except:
                                 supx = pd.DataFrame(columns=['AGE_BANDS', '12'])

                            df1 = pd.merge(df1,supx, on='AGE_BANDS', how='left')
                            df1['12'] = df1['12'].fillna(0)
                            # df1['5']

                            df1['6'] = df1['2'] + df1['4'] - df1['5']

                            df1['13'] = np.where(df1['6'] == 0,0,np.floor(((df1['11'] / df1['6']) * 100) + 0.5)).astype(int)
                            df1['1'] = df1['1'].astype(str)
                            df1['1'] = df1['1'].str.replace('10-12', onyr)
                            df1 = df1[['AGE_BANDS', '1', '2', '3', '4', '5','6', '7', '8', '9','10','11', '12', '13']].copy()
                            df1 = df1.set_index('AGE_BANDS')                                            
                                      
                        else:
                             df1 = dfdar.copy()
######################################################24 MONTHS
                        if t24mths.shape[0]>1:
                            t24mths[['Ayear', 'Amonth']] = t24mths[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                            
                            new = t24mths[((t24mths['Ayear']== year24) & (t24mths['Amonth'].isin(qmonths)))].copy()
                        else:
                             df2b = dfdar.copy()    
                        if new.shape[0]>1:
                            new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                            
                            orig4 = new[new['Tiyear']==994].copy()
                            
                            if orig4.shape[0]>0:
                                orig4['CD'] = pd.to_numeric(orig4['CD'], errors='coerce')
                                #  orig4a = orig4[['AGE_BANDS', 'ART']].copy()
                                orig4a = orig4.groupby('AGE_BANDS')['ART'].size().reset_index()
                                orig4a = orig4a.rename(columns={'ART':'2'})
                                orig4l = orig4[orig4['CD']<200].copy()
                                if orig4l.shape[0]>0:
                                        orig4l = orig4l.groupby('AGE_BANDS')['ART'].size().reset_index()
                                        orig4l = orig4l.rename(columns={'ART':'3'})
                                else:
                                    orig4l = datx.rename(columns={'1':'3'})         
                            else:
                                 orig4a = datx.rename(columns={'1':'2'})
                                 orig4l = datx.rename(columns={'1':'3'})
                            
                            tin = new[new['Tiyear']!=994].copy()
                            
                            if tin.shape[0]>0:
                                 tin4 = tin.groupby('AGE_BANDS')['ART'].size().reset_index()
                                 tin4 = tin4.rename(columns={'ART': '4'})
                            else:
                                 tin4 = datx.rename(columns ={'1':'4'})
                            
                            #one =new.shape[0]                        
                            #LOSSES
                                #LOSSES
                            new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                            newdead = new[new['Dyear']!=994].copy()
                            
                            new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD
                            #activ4E 1 YEAR
                            new['A'] = pd.to_numeric(new['A'], errors='coerce')
                            dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                            newcur = new[new['A'].isin(dfcur['A'])].copy()
                            
                            newlost = new[~new['A'].isin(dfcur['A'])].copy()
                            newlost['Tyear'] = pd.to_numeric(newlost['Tyear'], errors='coerce')
                            newlosto = newlost[newlost['Tyear']!=994].copy()
            
                            newlost = newlost[newlost['Tyear']==994].copy()
                            
                            if newlosto.shape[0] > 0:
                                    out24 = newlosto.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    out24 = out24.rename(columns={'ART':'5'})
                                    dato = daty['AGE_BANDS']
                                    out24 = pd.merge(dato, out24, on ='AGE_BANDS', how='left')
                                    out24['5'] = out['5'].fillna(0)
                                    
                            else:
                                    out24 = datx.rename(columns={'1':'5'})
                            
                            stop4 = datx.rename(columns = {'1':'7'})

                            if newdead.shape[0] > 0:
                                    dead24 = newdead.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    dead24 = dead24.rename(columns={'ART':'8'})           
                            else:
                                    dead24 = datx.rename(columns={'1':'8'})

                            miss4 = datx.rename(columns = {'1':'9'})

                            if newlost.shape[0] > 0:
                                    lost4 = newlost.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    lost4 = lost4.rename(columns={'ART':'10'})
                                    
                            else:
                                    lost4 = datx.rename(columns={'1':'10'})

                            if newcur.shape[0] > 0:
                                    activ4 = newcur.groupby('AGE_BANDS')['ART'].size().reset_index()
                                    newcur['VR'] = pd.to_numeric(newcur['VR'], errors='coerce')
                                    sup24 = newcur[newcur['VR'] <1000].copy()
                                    activ4 = activ4.rename(columns={'ART':'11'})
                                    if sup24.shape[0] > 0:
                                         sup24 = sup24.groupby('AGE_BANDS')['ART'].size().reset_index()
                                         sup24 = sup24.rename(columns={'ART':'12'})
                                    else:
                                         sup24 = datx.rename(columns={'1':'12'})
                                    
                            else:
                                    newcur = datx.rename(columns={'1':'11'})
                            df2b = pd.merge(daty,orig4a, on='AGE_BANDS', how='left')
                            df2b['2'] = df2b['2'].fillna(0)

                            df2b = pd.merge(df2b,orig4l, on='AGE_BANDS', how='left')
                            df2b['3'] = df2b['3'].fillna(0)

                            df2b = pd.merge(df2b,tin4, on='AGE_BANDS', how='left')
                            df2b['4'] = df2b['4'].fillna(0)

                            df2b = pd.merge(df2b,out24, on='AGE_BANDS', how='left')
                            df2b['5'] = df2b['5'].fillna(0)

                            df2b = pd.merge(df2b,stop4, on='AGE_BANDS', how='left')
                            df2b['7'] = df2b['7'].fillna(0)

                            df2b = pd.merge(df2b,dead24, on='AGE_BANDS', how='left')
                            df2b['8'] = df2b['8'].fillna(0)

                            df2b = pd.merge(df2b,miss4, on='AGE_BANDS', how='left')
                            df2b['9'] = df2b['9'].fillna(0)

                            df2b = pd.merge(df2b,lost4, on='AGE_BANDS', how='left')
                            df2b['10'] = df2b['10'].fillna(0)

                            df2b = pd.merge(df2b,activ4, on='AGE_BANDS', how='left')
                            df2b['11'] = df2b['11'].fillna(0)

                            df2b = pd.merge(df2b,sup24, on='AGE_BANDS', how='left')
                            df2b['12'] = df2b['12'].fillna(0)

                            df2b['6'] = df2b['2'] + df2b['4'] - df2b['5']

                            df2b['13'] = np.where(df2b['6'] == 0,0,np.floor(((df2b['11'] / df2b['6']) * 100) + 0.5)).astype(int)
                            df2b['1'] = df2b['1'].astype(str)
                            df2b['1'] = df2b['1'].str.replace('10-12', onyr)
                            df2b = df2b[['AGE_BANDS', '1', '2', '3', '4', '5','6', '7', '8', '9','10','11', '12', '13']].copy()
                            df2b = df2b.set_index('AGE_BANDS')                                            
                                      
                        else:
                             df2b = dfdar.copy()



                        oneyear[['Ayear', 'Amonth']] = oneyear[['Ayear', 'Amonth']].apply(pd.to_numeric, errors = 'coerce')
                        new = oneyear[((oneyear['Ayear']== oyear) & (oneyear['Amonth'].isin(qmonths)))].copy()
                        newtotal = new.shape[0]
                        new[['Tiyear']] = new[['Tiyear']].apply(pd.to_numeric, errors = 'coerce')
                        tin = new[new['Tiyear']!=994].copy()
                        #one =new.shape[0]
                        newti = tin.shape[0]
                        orig = int(newtotal)-int(newti)
                        #LOSSES
                        new['Dyear'] = pd.to_numeric(new['Dyear'], errors='coerce')
                        newdead = new[new['Dyear']!=994].copy()
                        deadnew = newdead.shape[0]
                        new = new[new['Dyear']==994].copy() #AFTER REMOVING THE DEAD
                        #ACTIVE 1 YEAR
                        new['A'] = pd.to_numeric(new['A'], errors='coerce')
                        dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                        newcur = new[new['A'].isin(dfcur['A'])].copy()

                        dfone = newcur.copy()

                        dfhc30_elig = dfone[['ART', 'GD', 'AGE BAND','A' ]].copy()

                        dfone[['Vyear', 'Vmonth']] = dfone[['Vyear', 'Vmonth']].apply(pd.to_numeric, errors='coerce')
                        dfone = dfone[((dfone['Vyear']==cyear) & (dfone['Vmonth'].isin(qmonths)))].copy()

                        dfhc30_test = dfone[['ART', 'GD', 'AGE BAND','A' ]].copy()

                        dfone['VR']  = pd.to_numeric(dfone['VR'], errors='coerce')
                        dfone = dfone[dfone['VR']<1000].copy()

                        dfhc30_sup = dfone[['ART', 'GD', 'AGE BAND','A' ]].copy()

                        newlost = new[~new['A'].isin(dfcur['A'])].copy()
                        newlost['Tyear'] = pd.to_numeric(newlost['Tyear'], errors='coerce')
                        newlosto = newlost[newlost['Tyear']!=994].copy()
                        newlostout = newlosto.shape[0]
                        newlost = newlost[newlost['Tyear']==994].copy()
        
                 
        
                        # #TRANSFER OUTS
                        ##RTT
        
                        #RTT BY LAST ENCOUNTER to include only months in the reporting Q
                        dfold = dfrt.copy()
                        dfold['Lyear'] = pd.to_numeric(dfold['Lyear'], errors='coerce') 
                        dfRTT = dfold[dfold['Lyear']== cyear].copy() #ALTER
                        dfRTT['Lmonth'] = pd.to_numeric(dfRTT['Lmonth'], errors='coerce') 
                        dfRTT = dfRTT[dfRTT['Lmonth'].isin(qmonths)].copy() #ALTER
        
                        #BY FIRST ENCOUNTER, To remove those first encountered in the Q
                        dfRTT[['Fyear', 'Fmonth']] = dfRTT[['Fyear', 'Fmonth']].apply(pd.to_numeric, errors='coerce') 
                        dfRTT = dfRTT[((dfRTT['Fyear']<cyear) | ((dfRTT['Fyear']== cyear) & (dfRTT['Fmonth']< fmonth)))].copy()
        
                        #BY RDDATE1,  take those that fall in the previous reporting Quarter
                        dfRTT['Royear'] = pd.to_numeric(dfRTT['Royear'], errors='coerce') 
                        dfRTTa = dfRTT[dfRTT['Royear']<lyear].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb = dfRTT[dfRTT['Royear']==lyear].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb[['R1month', 'R1day']] = dfRTTb[['R1month', 'R1day']].apply(pd.to_numeric, errors='coerce')
                        dfRTTb = dfRTTb[((dfRTTb['R1month']<lmonth) | ((dfRTTb['R1month']==lmonth) & (dfRTTb['R1day']<lday)))].copy()
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
                        dfRTT = dfRTT.copy()
        
                        #BY RD DATE2,  take those that fall in the previous reporting Quarter
                        dfRTT['R2year'] = pd.to_numeric(dfRTT['R2year'], errors='coerce')
                        dfRTTa = dfRTT[dfRTT['R2year']<lyear].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb = dfRTT[dfRTT['R2year']==lyear].copy() #CHANGE TO 2025 NEXT Q
                        dfRTTb[['R2month', 'R2day']] = dfRTTb[['R2month', 'R2day']].apply(pd.to_numeric, errors='coerce')
                        dfRTTb = dfRTTb[((dfRTTb['R2month']< lmonth) | ((dfRTTb['R2month']==lmonth) & (dfRTTb['R2day']<lday)))].copy()
                        dfRTT = pd.concat([dfRTTa, dfRTTb])
                        dfRTT = dfRTT.copy()
                        dfRTT['A'] = pd.to_numeric(dfRTT['A'], errors='coerce')
                        dfcur['A'] = pd.to_numeric(dfcur['A'], errors='coerce')
                        rtt = dfRTT[dfRTT['A'].isin(dfcur['A'])].copy()
                        rtta = rtt.shape[0]  
        ################################################################################################################
        #LIN   # dat = dat[['ART No.', 'RETURN DATE',  'LAST ENCOUNTER', 'TPT STATUS','LIKELIHOOD']].copy() 
    #################################################

    if st.session_state.reader:# and st.session_state.df:
        
            # Create an in-memory BytesIO buffer
            output = io.BytesIO()
            mapper = {'0-4':1, '5-9':2, '10-14':3, '15-19':4, '20-24':5, '25-29':6, '30-39':7, '40-49':8,'50+':9}               
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                if dfhc01.shape[0]>0:
                        hc01 = dfhc01.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc01 = hc01.reindex(columns=['M', 'F'], fill_value=0)
                        hc01 = hc01.reset_index()
                        hc01 = pd.merge(dfcomp, hc01, on='AGE BAND', how='left')
                        hc01['M'] = hc01['M'].fillna(0)
                        hc01['F'] = hc01['F'].fillna(0)            
                        # hc01['AGE BAND'] = hc01['AGE BAND'].astype(str)
                        hc01['SORT'] = hc01['AGE BAND'].map(mapper)
                        hc01['SORT'] = pd.to_numeric(hc01['SORT'], errors='coerce')
                        hc01 = hc01.sort_values(by ='SORT')
                        hc01 = hc01.drop(columns='SORT')
                        hc01.to_excel(writer, sheet_name="HC01", index=False)
                        hc01.to_excel(writer, sheet_name="HC02", index=False)
                        hc01.to_excel(writer, sheet_name="HC05", index=False)
                        hc01.to_excel(writer, sheet_name="HC06", index=False)
                else:
                        dfdar.to_excel(writer, sheet_name="HC01", index=False)
                        dfdar.to_excel(writer, sheet_name="HC02", index=False)
                        dfdar.to_excel(writer, sheet_name="HC05", index=False)
                        dfdar.to_excel(writer, sheet_name="HC06", index=False)


                if dfhc07.shape[0]>0:
                        hc07 = dfhc07.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc07 = hc07.reindex(columns=['M', 'F'], fill_value=0)
                        hc07 = hc07.reset_index()
                        hc07 = pd.merge(dfcomp, hc07, on='AGE BAND', how='left')
                        hc07['M'] = hc07['M'].fillna(0)
                        hc07['F'] = hc07['F'].fillna(0) 
                        # hc07['AGE BAND'] = hc07['AGE BAND'].astype(str)
                        hc07['SORT'] = hc07['AGE BAND'].map(mapper)
                        hc07['SORT'] = pd.to_numeric(hc07['SORT'], errors='coerce')
                        hc07 = hc07.sort_values(by ='SORT')
                        hc07 = hc07.drop(columns='SORT')
                        hc07.to_excel(writer, sheet_name="HC07", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC07", index=False)

                if dfhc08.shape[0]>0:
                        hc08 = dfhc08.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc08 = hc08.reindex(columns=['M', 'F'], fill_value=0)
                        hc08 = hc08.reset_index()
                        hc08 = pd.merge(dfcomp, hc08, on='AGE BAND', how='left')
                        hc08['M'] = hc08['M'].fillna(0)
                        hc08['F'] = hc08['F'].fillna(0)
                        # hc08['AGE BAND'] = hc08['AGE BAND'].astype(str)
                        hc08['SORT'] = hc08['AGE BAND'].map(mapper)
                        hc08['SORT'] = pd.to_numeric(hc08['SORT'], errors='coerce')
                        hc08 = hc08.sort_values(by ='SORT')
                        hc08 = hc08.drop(columns='SORT')
                        hc08.to_excel(writer, sheet_name="HC08", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC08", index=False)   

                if dfhc09.shape[0]>0:
                        hc09 = dfhc09.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc09 = hc09.reindex(columns=['M', 'F'], fill_value=0)
                        hc09 = hc09.reset_index()
                        hc09 = pd.merge(dfcomp, hc09, on='AGE BAND', how='left')
                        hc09['M'] = hc09['M'].fillna(0)
                        hc09['F'] = hc09['F'].fillna(0)
                        # hc09['AGE BAND'] = hc09['AGE BAND'].astype(str)
                        hc09['SORT'] = hc09['AGE BAND'].map(mapper)
                        hc09['SORT'] = pd.to_numeric(hc09['SORT'], errors='coerce')
                        hc09 = hc09.sort_values(by ='SORT')
                        hc09 = hc09.drop(columns='SORT')
                        hc09.to_excel(writer, sheet_name="HC09", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC09", index=False)             

                if dfhc11a.shape[0]>0:
                        hc11a = dfhc11a.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc11a = hc11a.reindex(columns=['M', 'F'], fill_value=0)
                        hc11a = hc11a.reset_index()
                        hc11a = pd.merge(dfcomp, hc11a, on='AGE BAND', how='left')
                        hc11a['M'] = hc11a['M'].fillna(0)
                        hc11a['F'] = hc11a['F'].fillna(0)
                        # hc11a['AGE BAND'] = hc11a['AGE BAND'].astype(str)
                        hc11a['SORT'] = hc11a['AGE BAND'].map(mapper)
                        hc11a['SORT'] = pd.to_numeric(hc11a['SORT'], errors='coerce')
                        hc11a = hc11a.sort_values(by ='SORT')
                        hc11a = hc11a.drop(columns='SORT')
                        hc11a.to_excel(writer, sheet_name="HC11a", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC11a", index=False)

                if dfhc11b.shape[0]>0:
                        hc11b = dfhc11b.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc11b = hc11b.reindex(columns=['M', 'F'], fill_value=0)
                        hc11b = hc11b.reset_index()
                        hc11b = pd.merge(dfcomp, hc11b, on='AGE BAND', how='left')
                        hc11b['M'] = hc11b['M'].fillna(0)
                        hc11b['F'] = hc11b['F'].fillna(0) 
                        # hc11b['AGE BAND'] = hc11b['AGE BAND'].astype(str)
                        hc11b['SORT'] = hc11b['AGE BAND'].map(mapper)
                        hc11b['SORT'] = pd.to_numeric(hc11b['SORT'], errors='coerce')
                        hc11b = hc11b.sort_values(by ='SORT')
                        hc11b = hc11b.drop(columns='SORT')
                        hc11b.to_excel(writer, sheet_name="HC11b", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC11b", index=False)

                dfdar.to_excel(writer, sheet_name="HC11c", index=False)

                if dfhc11d.shape[0]>0:
                        hc11d = dfhc11d.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc11d = hc11d.reindex(columns=['M', 'F'], fill_value=0)
                        hc11d = hc11d.reset_index()
                        hc11d = pd.merge(dfcomp, hc11d, on='AGE BAND', how='left')
                        hc11d['M'] = hc11d['M'].fillna(0)
                        hc11d['F'] = hc11d['F'].fillna(0) 
                        # hc11d['AGE BAND'] = hc11d['AGE BAND'].astype(str)
                        hc11d['SORT'] = hc11d['AGE BAND'].map(mapper)
                        hc11d['SORT'] = pd.to_numeric(hc11d['SORT'], errors='coerce')
                        hc11d = hc11d.sort_values(by ='SORT')
                        hc11d = hc11d.drop(columns='SORT')
                        hc11d.to_excel(writer, sheet_name="HC11d", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC11d", index=False)

                if dfhc12_3.shape[0]>0:
                        hc12_3 = dfhc12_3.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc12_3 = hc12_3.reindex(columns=['M', 'F'], fill_value=0)
                        hc12_3 = hc12_3.reset_index()
                        hc12_3 = pd.merge(dfcomp, hc12_3, on='AGE BAND', how='left')
                        hc12_3['M'] = hc12_3['M'].fillna(0)
                        hc12_3['F'] = hc12_3['F'].fillna(0) 
                        # hc12_3['AGE BAND'] = hc12_3['AGE BAND'].astype(str)
                        hc12_3['SORT'] = hc12_3['AGE BAND'].map(mapper)
                        hc12_3['SORT'] = pd.to_numeric(hc12_3['SORT'], errors='coerce')
                        hc12_3 = hc12_3.sort_values(by ='SORT')
                        hc12_3 = hc12_3.drop(columns='SORT')
                        hc12_3.to_excel(writer, sheet_name="HC12<3", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC12<3", index=False)

                if dfhc12_5.shape[0]>0:
                        hc12_5 = dfhc12_5.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc12_5 = hc12_5.reindex(columns=['M', 'F'], fill_value=0)
                        hc12_5 = hc12_5.reset_index()
                        hc12_5 = pd.merge(dfcomp, hc12_5, on='AGE BAND', how='left')
                        hc12_5['M'] = hc12_5['M'].fillna(0)
                        hc12_5['F'] = hc12_5['F'].fillna(0) 
                        # hc12_5['AGE BAND'] = hc12_5['AGE BAND'].astype(str)
                        hc12_5['SORT'] = hc12_5['AGE BAND'].map(mapper)
                        hc12_5['SORT'] = pd.to_numeric(hc12_5['SORT'], errors='coerce')
                        hc12_5 = hc12_5.sort_values(by ='SORT')
                        hc12_5 = hc12_5.drop(columns='SORT')
                        hc12_5.to_excel(writer, sheet_name="HC12_5", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC12_5", index=False)

                if dfhc12_6.shape[0]>0:
                        hc12_6 = dfhc12_6.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc12_6 = hc12_6.reindex(columns=['M', 'F'], fill_value=0)
                        hc12_6 = hc12_6.reset_index()
                        hc12_6 = pd.merge(dfcomp, hc12_6, on='AGE BAND', how='left')
                        hc12_6['M'] = hc12_6['M'].fillna(0)
                        hc12_6['F'] = hc12_6['F'].fillna(0) 
                        # hc12_6['AGE BAND'] = hc12_6['AGE BAND'].astype(str)
                        hc12_6['SORT'] = hc12_6['AGE BAND'].map(mapper)
                        hc12_6['SORT'] = pd.to_numeric(hc12_6['SORT'], errors='coerce')
                        hc12_6 = hc12_6.sort_values(by ='SORT')
                        hc12_6 = hc12_6.drop(columns='SORT')
                        hc12_6.to_excel(writer, sheet_name="HC12_6", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC12_6", index=False)

                if dfhc14_TO.shape[0]>0:
                        hc14_TO = dfhc14_TO.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc14_TO = hc14_TO.reindex(columns=['M', 'F'], fill_value=0)
                        hc14_TO = hc14_TO.reset_index()
                        hc14_TO = pd.merge(dfcomp, hc14_TO, on='AGE BAND', how='left')
                        hc14_TO['M'] = hc14_TO['M'].fillna(0)
                        hc14_TO['F'] = hc14_TO['F'].fillna(0) 
                        # hc14_TO['AGE BAND'] = hc14_TO['AGE BAND'].astype(str)
                        hc14_TO['SORT'] = hc14_TO['AGE BAND'].map(mapper)
                        hc14_TO['SORT'] = pd.to_numeric(hc14_TO['SORT'], errors='coerce')
                        hc14_TO = hc14_TO.sort_values(by ='SORT')
                        hc14_TO = hc14_TO.drop(columns='SORT')
                        hc14_TO.to_excel(writer, sheet_name="HC14_TO", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC14_TO", index=False)

                dfdar.to_excel(writer, sheet_name="HC14_STOP", index=False)

                if dfhc14_LOST.shape[0]>0:
                        hc14_LOST = dfhc14_LOST.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc14_LOST = hc14_LOST.reindex(columns=['M', 'F'], fill_value=0)
                        hc14_LOST = hc14_LOST.reset_index()
                        hc14_LOST = pd.merge(dfcomp, hc14_LOST, on='AGE BAND', how='left')
                        hc14_LOST['M'] = hc14_LOST['M'].fillna(0)
                        hc14_LOST['F'] = hc14_LOST['F'].fillna(0) 
                        # hc14_LOST['AGE BAND'] = hc14_LOST['AGE BAND'].astype(str)
                        hc14_LOST['SORT'] = hc14_LOST['AGE BAND'].map(mapper)
                        hc14_LOST['SORT'] = pd.to_numeric(hc14_LOST['SORT'], errors='coerce')
                        hc14_LOST = hc14_LOST.sort_values(by ='SORT')
                        hc14_LOST = hc14_LOST.drop(columns='SORT')
                        hc14_LOST.to_excel(writer, sheet_name="HC14_LOST", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC14_LOST", index=False)

                if dfhc14_DIED.shape[0]>0:
                        hc14_DIED = dfhc14_DIED.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc14_DIED = hc14_DIED.reindex(columns=['M', 'F'], fill_value=0)
                        hc14_DIED = hc14_DIED.reset_index()
                        hc14_DIED = pd.merge(dfcomp, hc14_DIED, on='AGE BAND', how='left')
                        hc14_DIED['M'] = hc14_DIED['M'].fillna(0)
                        hc14_DIED['F'] = hc14_DIED['F'].fillna(0) 
                        # hc14_DIED['AGE BAND'] = hc14_DIED['AGE BAND'].astype(str)
                        hc14_DIED['SORT'] = hc14_DIED['AGE BAND'].map(mapper)
                        hc14_DIED['SORT'] = pd.to_numeric(hc14_DIED['SORT'], errors='coerce')
                        hc14_DIED = hc14_DIED.sort_values(by ='SORT')
                        hc14_DIED = hc14_DIED.drop(columns='SORT')
                        hc14_DIED.to_excel(writer, sheet_name="HC14_DIED", index=False)
                        hc14_DIED.to_excel(writer, sheet_name="HC15", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC14_DIED", index=False)
                   dfdar.to_excel(writer, sheet_name="HC15", index=False)

                if dfhc01.shape[0]>0:
                        hc01.to_excel(writer, sheet_name="HC18_scr", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC18_scr", index=False)

                if dfhc19_scr.shape[0]>0:
                        hc19_scr = dfhc19_scr.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc19_scr = hc19_scr.reindex(columns=['M', 'F'], fill_value=0)
                        hc19_scr = hc19_scr.reset_index()
                        hc19_scr = pd.merge(dfcomp, hc19_scr, on='AGE BAND', how='left')
                        hc19_scr['M'] = hc19_scr['M'].fillna(0)
                        hc19_scr['F'] = hc19_scr['F'].fillna(0) 
                        # hc19_scr['AGE BAND'] = hc19_scr['AGE BAND'].astype(str)
                        hc19_scr['SORT'] = hc19_scr['AGE BAND'].map(mapper)
                        hc19_scr['SORT'] = pd.to_numeric(hc19_scr['SORT'], errors='coerce')
                        hc19_scr = hc19_scr.sort_values(by ='SORT')
                        hc19_scr = hc19_scr.drop(columns='SORT')
                        hc19_scr.to_excel(writer, sheet_name="HC19_scr", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC19_scr", index=False)

                if dfhc18_scr.shape[0]>0:
                        hc18_scr = dfhc18_scr.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc18_scr = hc18_scr.reindex(columns=['M', 'F'], fill_value=0)
                        hc18_scr = hc18_scr.reset_index()
                        hc18_scr = pd.merge(dfcomp, hc18_scr, on='AGE BAND', how='left')
                        hc18_scr['M'] = hc18_scr['M'].fillna(0)
                        hc18_scr['F'] = hc18_scr['F'].fillna(0) 
                        # hc18_scr['AGE BAND'] = hc18_scr['AGE BAND'].astype(str)
                        hc18_scr['SORT'] = hc18_scr['AGE BAND'].map(mapper)
                        hc18_scr['SORT'] = pd.to_numeric(hc18_scr['SORT'], errors='coerce')
                        hc18_scr = hc18_scr.sort_values(by ='SORT')
                        hc18_scr = hc18_scr.drop(columns='SORT')
                        hc18_scr.to_excel(writer, sheet_name="HC23_scr", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC23_scr", index=False)

                if dfhc26_Elig.shape[0]>0:
                        hc26_Elig = dfhc26_Elig.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc26_Elig = hc26_Elig.reindex(columns=['M', 'F'], fill_value=0)
                        hc26_Elig = hc26_Elig.reset_index()
                        hc26_Elig = pd.merge(dfcomp, hc26_Elig, on='AGE BAND', how='left')
                        hc26_Elig['M'] = hc26_Elig['M'].fillna(0)
                        hc26_Elig['F'] = hc26_Elig['F'].fillna(0) 
                        # hc26_Elig['AGE BAND'] = hc26_Elig['AGE BAND'].astype(str)
                        hc26_Elig['SORT'] = hc26_Elig['AGE BAND'].map(mapper)
                        hc26_Elig['SORT'] = pd.to_numeric(hc26_Elig['SORT'], errors='coerce')
                        hc26_Elig = hc26_Elig.sort_values(by ='SORT')
                        hc26_Elig = hc26_Elig.drop(columns='SORT')
                        hc26_Elig.to_excel(writer, sheet_name="HC26_Elig", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC26_Elig", index=False)
          
                if dfhc26_test.shape[0]>0:
                        hc26_test = dfhc26_test.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc26_test = hc26_test.reindex(columns=['M', 'F'], fill_value=0)
                        hc26_test = hc26_test.reset_index()
                        hc26_test = pd.merge(dfcomp, hc26_test, on='AGE BAND', how='left')
                        hc26_test['M'] = hc26_test['M'].fillna(0)
                        hc26_test['F'] = hc26_test['F'].fillna(0) 
                        # hc26_test['AGE BAND'] = hc26_test['AGE BAND'].astype(str)
                        hc26_test['SORT'] = hc26_test['AGE BAND'].map(mapper)
                        hc26_test['SORT'] = pd.to_numeric(hc26_test['SORT'], errors='coerce')
                        hc26_test = hc26_test.sort_values(by ='SORT')
                        hc26_test = hc26_test.drop(columns='SORT')
                        hc26_test.to_excel(writer, sheet_name="HC26_test", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC26_test", index=False)

                if dfhc26_sup.shape[0]>0:
                        hc26_sup = dfhc26_sup.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc26_sup = hc26_sup.reindex(columns=['M', 'F'], fill_value=0)
                        hc26_sup = hc26_sup.reset_index()
                        hc26_sup = pd.merge(dfcomp, hc26_sup, on='AGE BAND', how='left')
                        hc26_sup['M'] = hc26_sup['M'].fillna(0)
                        hc26_sup['F'] = hc26_sup['F'].fillna(0) 
                        # hc26_sup['AGE BAND'] = hc26_sup['AGE BAND'].astype(str)
                        hc26_sup['SORT'] = hc26_sup['AGE BAND'].map(mapper)
                        hc26_sup['SORT'] = pd.to_numeric(hc26_sup['SORT'], errors='coerce')
                        hc26_sup = hc26_sup.sort_values(by ='SORT')
                        hc26_sup = hc26_sup.drop(columns='SORT')
                        hc26_sup.to_excel(writer, sheet_name="HC26_sup", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC26_sup", index=False)

                if dfhc29_elig.shape[0]>0:
                         hc29_elig = dfhc29_elig.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                         hc29_elig =  hc29_elig.reindex(columns=['M', 'F'], fill_value=0)
                         hc29_elig =  hc29_elig.reset_index()
                         hc29_elig = pd.merge(dfcomp, hc29_elig, on='AGE BAND', how='left')
                         hc29_elig['M'] = hc29_elig['M'].fillna(0)
                         hc29_elig['F'] = hc29_elig['F'].fillna(0) 
                        #  hc29_elig['AGE BAND'] =  hc29_elig['AGE BAND'].astype(str)
                         hc29_elig['SORT'] =  hc29_elig['AGE BAND'].map(mapper)
                         hc29_elig['SORT'] = pd.to_numeric( hc29_elig['SORT'], errors='coerce')
                         hc29_elig =  hc29_elig.sort_values(by ='SORT')
                         hc29_elig =  hc29_elig.drop(columns='SORT')
                         hc29_elig.to_excel(writer, sheet_name="HC29_ELIG", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC29_ELIG", index=False)

                if dfhc29_test.shape[0]>0:
                         hc29_test = dfhc29_test.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                         hc29_test =  hc29_test.reindex(columns=['M', 'F'], fill_value=0)
                         hc29_test =  hc29_test.reset_index()
                         hc29_test = pd.merge(dfcomp, hc29_test, on='AGE BAND', how='left')
                         hc29_test['M'] = hc29_test['M'].fillna(0)
                         hc29_test['F'] = hc29_test['F'].fillna(0) 
                        #  hc29_test['AGE BAND'] =  hc29_test['AGE BAND'].astype(str)
                         hc29_test['SORT'] =  hc29_test['AGE BAND'].map(mapper)
                         hc29_test['SORT'] = pd.to_numeric( hc29_test['SORT'], errors='coerce')
                         hc29_test =  hc29_test.sort_values(by ='SORT')
                         hc29_test =  hc29_test.drop(columns='SORT')
                         hc29_test.to_excel(writer, sheet_name="HC29_TEST", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC29_TEST", index=False)


                if dfhc29_sup.shape[0]>0:
                         hc29_sup = dfhc29_sup.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                         hc29_sup =  hc29_sup.reindex(columns=['M', 'F'], fill_value=0)
                         hc29_sup =  hc29_sup.reset_index()
                         hc29_sup = pd.merge(dfcomp, hc29_sup, on='AGE BAND', how='left')
                         hc29_sup['M'] = hc29_sup['M'].fillna(0)
                         hc29_sup['F'] = hc29_sup['F'].fillna(0) 
                        #  hc29_sup['AGE BAND'] =  hc29_sup['AGE BAND'].astype(str)
                         hc29_sup['SORT'] =  hc29_sup['AGE BAND'].map(mapper)
                         hc29_sup['SORT'] = pd.to_numeric( hc29_sup['SORT'], errors='coerce')
                         hc29_sup =  hc29_sup.sort_values(by ='SORT')
                         hc29_sup =  hc29_sup.drop(columns='SORT')
                         hc29_sup.to_excel(writer, sheet_name="HC29_SUP", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC29_SUP", index=False)

                if dfhc30_elig.shape[0]>0:
                        hc30_elig = dfhc30_elig.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc30_elig = hc30_elig.reindex(columns=['M', 'F'], fill_value=0)
                        hc30_elig = hc30_elig.reset_index()
                        hc30_elig = pd.merge(dfcomp, hc30_elig, on='AGE BAND', how='left')
                        hc30_elig['M'] = hc30_elig['M'].fillna(0)
                        hc30_elig['F'] = hc30_elig['F'].fillna(0) 
                        # hc30_elig['AGE BAND'] = hc30_elig['AGE BAND'].astype(str)
                        hc30_elig['SORT'] = hc30_elig['AGE BAND'].map(mapper)
                        hc30_elig['SORT'] = pd.to_numeric(hc30_elig['SORT'], errors='coerce')
                        hc30_elig = hc30_elig.sort_values(by ='SORT')
                        hc30_elig = hc30_elig.drop(columns='SORT')
                        hc30_elig.to_excel(writer, sheet_name="HC30_ELIG", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC30_ELIG", index=False)

                if dfhc30_test.shape[0]>0:
                        hc30_test = dfhc30_test.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc30_test = hc30_test.reindex(columns=['M', 'F'], fill_value=0)
                        hc30_test = hc30_test.reset_index()
                        hc30_test = pd.merge(dfcomp, hc30_test, on='AGE BAND', how='left')
                        hc30_test['M'] = hc30_test['M'].fillna(0)
                        hc30_test['F'] = hc30_test['F'].fillna(0) 
                        # hc30_test['AGE BAND'] = hc30_test['AGE BAND'].astype(str)
                        hc30_test['SORT'] = hc30_test['AGE BAND'].map(mapper)
                        hc30_test['SORT'] = pd.to_numeric(hc30_test['SORT'], errors='coerce')
                        hc30_test = hc30_test.sort_values(by ='SORT')
                        hc30_test = hc30_test.drop(columns='SORT')
                        hc30_test.to_excel(writer, sheet_name="HC30_TEST", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC30_TEST", index=False)


                if dfhc30_sup.shape[0]>0:
                        hc30_sup = dfhc30_sup.groupby(['AGE BAND', 'GD']).size().unstack(fill_value=0)
                        hc30_sup = hc30_sup.reindex(columns=['M', 'F'], fill_value=0)
                        hc30_sup = hc30_sup.reset_index()
                        hc30_sup = pd.merge(dfcomp, hc30_sup, on='AGE BAND', how='left')
                        hc30_sup['M'] = hc30_sup['M'].fillna(0)
                        hc30_sup['F'] = hc30_sup['F'].fillna(0) 
                        # hc30_sup['AGE BAND'] = hc30_sup['AGE BAND'].astype(str)
                        hc30_sup['SORT'] = hc30_sup['AGE BAND'].map(mapper)
                        hc30_sup['SORT'] = pd.to_numeric(hc30_sup['SORT'], errors='coerce')
                        hc30_sup = hc30_sup.sort_values(by ='SORT')
                        hc30_sup = hc30_sup.drop(columns='SORT')
                        hc30_sup.to_excel(writer, sheet_name="HC30_SUP", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC30_SUP", index=False)

                if dfhc43 .shape[0]>0:
                        hc43  = dfhc43 .groupby(['DSDM','AGE BAND','GD']).size().unstack(fill_value=0)
                        hc43  = hc43 .reindex(columns=['M', 'F'], fill_value=0)
                        hc43  = hc43 .reset_index()
                        # hc43 ['AGE BAND'] = hc43 ['AGE BAND'].astype(str)
                        hc43 ['SORT'] = hc43 ['AGE BAND'].map(mapper)
                        hc43 ['SORT'] = pd.to_numeric(hc43 ['SORT'], errors='coerce')
                        hc43  = hc43 .sort_values(by =['DSDM', 'SORT'])
                        hc43  = hc43 .drop(columns='SORT')
                        hc43 .to_excel(writer, sheet_name="HC43", index=False)
                else:
                   dfdar.to_excel(writer, sheet_name="HC43", index=False)

                if sixmths.shape[0]>0:

                        df2a .to_excel(writer, sheet_name="CA01a", index=True)
                else:
                   dfdar.to_excel(writer, sheet_name="CA01a", index=True)
            
                if oneyear.shape[0]>0:

                        df1 .to_excel(writer, sheet_name="CA02a", index=True)
                else:
                   dfdar.to_excel(writer, sheet_name="CA02a", index=True)

                if t24mths.shape[0]>0:

                        df2b.to_excel(writer, sheet_name="CA03a", index=True)
                else:
                   dfdar.to_excel(writer, sheet_name="CA03a", index= True)



    
                output.seek(0)    
                            # Provide one combined download button
            facility = st.text_input('FACILITY NAME')
            if not facility:
                             st.stop()
            else:
                            st.download_button(
                                label="📥 DOWNLOAD PIVOT TABLES",
                                data=output,
                                file_name=f"{facility}_106a_PIVOTS.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
     

                            st.success('**CREATED BY Dr. LUMINSA DESIRE**')
                        
pages = {
    "READER:": [
        st.Page(extract, title="EMR EXTRACT READER"),
    ],
   
}

pg = st.navigation(pages)
pg.run()
                                
    

