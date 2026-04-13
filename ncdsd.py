import pandas as pd
import streamlit as st 
import os
import io
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import traceback
import time
# from google.oauth2.service_account import Credentials
# from oauth2client.service_account import ServiceAccountCredentials
# from streamlit_gsheets import GSheetsConnection
from datetime import datetime 
import datetime as dt
st.cache_data.clear()
st.cache_resource.clear()

def extract():
    cola,colb,colc = st.columns([1,3,1])
    colb.subheader('NCDS DHIS2')   
    file = st.file_uploader("Upload your EMR extract here", type=['xlsx']) 
    if file is not None:   
        if 'fd' not in st.session_state:
            fileN = file.name
            name = os.path.basename(fileN).split('.')[0]
            st.session_state.fd = name
        else:
            pass
    else:
        pass
    if file is not None: 
       fileN = file.name
       namey = os.path.basename(fileN).split('.')[0]
       if str(namey) != str(st.session_state.fd):
                #st.info(f'DATA FOR {facy} NOT SUBMITTED')
                st.session_state.submited = False
                st.cache_data.clear()
                st.session_state.fd = namey
                st.cache_resource.clear()
                st.session_state.submited =False
                st.session_state.dfw = None
                st.session_state.readery =False#
                time.sleep(1)
                st.rerun()         
    if 'submited' not in st.session_state:
        st.session_state.submited =False
    if 'df' not in st.session_state:
        st.session_state.dfw = None
    # if 'fac' not in st.session_state:
   # st.session_state.fac = None
    if 'reader' not in st.session_state:
        st.session_state.readery =False#
    #ext = None
    if file is not None and not st.session_state.readery:
        # Get the file name
        fileN = file.name
        ext = os.path.basename(fileN).split('.')[1]
        # if ext == 'xlsx.xlsx':
        #        ext = 'xlsx'
    #df = None
    if file is not None and not st.session_state.readery:
        wb = load_workbook(file)
        sheets = wb.sheetnames
        if len(sheets)>1:
            st. warning('THIS EXTRACT HAS MULTIPLE SHEETS, I CAN NOT TELL WHICH ONE TO READ')
            time.sleep(3)
            st.info('DELETE ALL THE OTHER SHEETS AND REMAIN WITH ONE THAT HAS THE EVER ENROLLED')
            st.stop()
        else:
            pass

    if file is not None and not st.session_state.readery:
                    st.session_state.dfw = pd.read_excel(file)
                    df = st.session_state.dfw 
    #if file is not None and not st.session_state.rea
                    df = df.rename(columns= {'ART  ':'ART',  'RD  ':'RD', 'LD  ': 'LD',
       'TI  ': 'TI', 'TO  ':'TO', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD'})
                    df = df.rename(columns= {'ART ':'ART', 'RD ':'RD', 'LD ': 'LD',
                            'TI ': 'TI', 'TO ':'TO', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD'})
                    columns = ['ART','AG', 'GD', 'RD','TO', 'TI', 'DD','LD']
                    cols = df.columns.to_list()
                    if not all(column in cols for column in columns):
                        missing_columns = [column for column in columns if column not in cols]
                        for column in missing_columns:
                            st.markdown(f' **ERROR !!! {column} is not in the file uploaded**')
                            st.markdown('**First rename all the columns as guided above**')
                            st.stop()
                    st.session_state.readery= True
    if st.session_state.readery:
                          # Convert 'ART' column to string and create 'ART' column with numeric part to remove blanks
                        st.session_state.dfw = st.session_state.dfw.rename(columns= {'ART  ':'ART', 'RD  ':'RD', 'LD  ': 'LD',
        'TI  ': 'TI', 'TO  ':'TO', 'DD  ': 'DD', 'AG  ':'AG', 'GD  ':'GD'})
                        st.session_state.dfw = st.session_state.dfw.rename(columns= {'ART ':'ART',  'RD ':'RD',  'LD ': 'LD',
                            'TI ': 'TI', 'TO ':'TO', 'DD ': 'DD', 'AG ':'AG', 'GD ':'GD'})
                        df = st.session_state.dfw.copy()
                        
                        df = st.session_state.dfw[['ART', 'AG', 'RD','GD','TO', 'TI', 'DD','LD','HTN','DM','MH']].copy()
                        df['ART'] = df['ART'].astype(str)
                        df['A'] = df['ART'].str.replace('[^0-9]', '', regex=True)
                        df['A'] = pd.to_numeric(df['A'], errors= 'coerce')
                        df = df[df['A']>0].copy()
                        #df.dropna(subset='ART', inplace=True)                  
                        df[[ 'RD','TO','TI']] = df[[ 'RD','TO','TI']].astype(str)
                        if df['TI'].str.contains('YES').any():
                            st.write("You may be using the Transfer in column instead of the Transfer_in Obs date column")
                            st.stop()
                       
                        
                        df['RD'] = df['RD'].astype(str)
                        
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                                        
                        y = pd.DataFrame({'ART' :['2','3','4','5'], 'TI':['1-1-1',1,'1/1/1','3 8 2001'], 'RD':['1-1-1',1,'1/1/1','3 8 2001'],'DD':['1-1-1',1,'1/1/1','3 8 2001'], 
                                        'TO':['1-1-1',1,'1/1/1','3 8 2001'],
                                              'LD':['1-1-1',1,'1/1/1','3 8 2001']})                        
                        
                       
                        df['RD'] = df['RD'].astype(str)
                      
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                        
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        
                        df['RD'] = df['RD'].str.replace('00:00:00', '', regex=True)
                   
                        df['TI'] = df['TI'].str.replace('00:00:00', '', regex=True)
                        df['TO'] = df['TO'].str.replace('00:00:00', '', regex=True)
                       
                        df['DD'] = df['DD'].str.replace('00:00:00', '', regex=True)
                        df['LD'] = df['LD'].str.replace('00:00:00', '', regex=True)
                        
                        df["TI"] = df["TI"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["DD"] = df["DD"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        df["RD"] = df["RD"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()

                        
                        df["TO"] = df["TO"].str.replace(r"\s*\d{1,2}:\d{2}.*$", "", regex=True).str.strip()
                        
                        df["TI"] = df["TI"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["DD"] = df["DD"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        df["RD"] = df["RD"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
     
                        
                        df["TO"] = df["TO"].str.replace(r"\s*\..*$", "", regex=True).str.strip()
                        
                        df = pd.concat([df,y])
                        df = df.copy()
                       
                        
                        df['RD'] = df['RD'].astype(str) ###
                 #
                        df['TI'] = df['TI'].astype(str) ##
                        df['TO'] = df['TO'].astype(str) ##
                        
                        df['DD'] = df['DD'].astype(str) ####
                        df['LD'] = df['LD'].astype(str)
                                

                        # SPLITTING DEATH DATE
                        A = df[df['DD'].str.contains('-')].copy()
                        a = df[~df['DD'].str.contains('-')].copy()
                        B = a[a['DD'].str.contains('/')].copy()
                        C = a[~a['DD'].str.contains('/')].copy()
                        E = C[C['DD'].str.contains(' ')].copy()
                        D = C[~C['DD'].str.contains(' ')].copy()
                        A[['Dyear', 'Dmonth', 'Dday']] = A['DD'].str.split('-', expand = True)
                        B[['Dyear', 'Dmonth', 'Dday']] = B['DD'].str.split('/', expand = True)
                        try:
                            D['DD'] = pd.to_numeric(D['DD'], errors='coerce')
                            D['DD'] = pd.to_datetime(D['DD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['DD'] =  D['DD'].astype(str)
                            D[['Dyear', 'Dmonth', 'Dday']] = D['DD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['DD'] = pd.to_datetime(E['DD'],format='%d %m %Y', errors='coerce')
                            E['DD'] =  E['DD'].astype(str)
                            E[['Dyear', 'Dmonth', 'Dday']] = E['DD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])                     
                        # SORTING THE RETURN VISIT DATE
                        A = df[df['RD'].str.contains('-')].copy()
                        a = df[~df['RD'].str.contains('-')].copy()
                        B = a[a['RD'].str.contains('/')].copy()
                        C = a[~a['RD'].str.contains('/')].copy()
                        E = C[C['RD'].str.contains(' ')].copy()
                        D = C[~C['RD'].str.contains(' ')].copy()                     
                        #D = C[C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()
                        #E = C[~C['RD'].apply(lambda x: isinstance(x, (int, float)) or x.isdigit())].copy()              
                        
                        A[['Ryear', 'Rmonth', 'Rday']] = A['RD'].str.split('-', expand = True)
                        B[['Ryear', 'Rmonth', 'Rday']] = B['RD'].str.split('/', expand = True)
                        try:
                            D['RD'] = pd.to_numeric(D['RD'], errors='coerce')
                            D['RD'] = pd.to_datetime(D['RD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['RD'] =  D['RD'].astype(str)
                            D[['Ryear', 'Rmonth', 'Rday']] = D['RD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['RD'] = pd.to_datetime(E['RD'],format='%d %m %Y', errors='coerce')
                            E['RD'] =  E['RD'].astype(str)
                            E[['Ryear', 'Rmonth', 'Rday']] = E['RD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])                
                        
                        #SORTING THE TO DATE
                        A = df[df['TO'].str.contains('-')].copy()
                        a = df[~df['TO'].str.contains('-')].copy()
                        B = a[a['TO'].str.contains('/')].copy()
                        C = a[~a['TO'].str.contains('/')].copy()
                        E = C[C['TO'].str.contains(' ')].copy()
                        D = C[~C['TO'].str.contains(' ')].copy()         
                        A[['Tyear', 'Tmonth', 'Tday']] = A['TO'].str.split('-', expand = True)
                        B[['Tyear', 'Tmonth', 'Tday']] = B['TO'].str.split('/', expand = True)
                        try:
                            D['TO'] = pd.to_numeric(D['TO'], errors='coerce')
                            D['TO'] = pd.to_datetime(D['TO'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TO'] =  D['TO'].astype(str)
                            D[['Tyear', 'Tmonth', 'Tday']] = D['TO'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TO'] = pd.to_datetime(E['TO'],format='%d %m %Y', errors='coerce')
                            E['TO'] =  E['TO'].astype(str)
                            E[['Tyear', 'Tmonth', 'Tday']] = E['TO'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])         
                    #SORTING THE TI DATE
                        A = df[df['TI'].str.contains('-')].copy()
                        a = df[~df['TI'].str.contains('-')].copy()
                        B = a[a['TI'].str.contains('/')].copy()
                        C = a[~a['TI'].str.contains('/')].copy()
                        E = C[C['TI'].str.contains(' ')].copy()
                        D = C[~C['TI'].str.contains(' ')].copy()         
                        A[['Tiyear', 'Timonth', 'Tiday']] = A['TI'].str.split('-', expand = True)
                        B[['Tiyear', 'Timonth', 'Tiday']] = B['TI'].str.split('/', expand = True)
                        try:
                            D['TI'] = pd.to_numeric(D['TI'], errors='coerce')
                            D['TI'] = pd.to_datetime(D['TI'], origin='1899-12-30', unit='D', errors='coerce')
                            D['TI'] =  D['TI'].astype(str)
                            D[['Tiyear', 'Timonth', 'Tiday']] = D['TI'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['TI'] = pd.to_datetime(E['TI'],format='%d %m %Y', errors='coerce')
                            E['TI'] =  E['TI'].astype(str)
                            E[['Tiyear', 'Timonth', 'Tiday']] = E['TI'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                       
                        # SORTING THE LAST ENCOUNTER DATES
                        A = df[df['LD'].str.contains('-')].copy()
                        a = df[~df['LD'].str.contains('-')].copy()
                        B = a[a['LD'].str.contains('/')].copy()
                        C = a[~a['LD'].str.contains('/')].copy()
                        E = C[C['LD'].str.contains(' ')].copy()
                        D = C[~C['LD'].str.contains(' ')].copy()
                        A[['Lyear', 'Lmonth', 'Lday']] = A['LD'].str.split('-', expand = True)
                        B[['Lyear', 'Lmonth', 'Lday']] = B['LD'].str.split('/', expand = True)
                        try:
                            D['LD'] = pd.to_numeric(D['LD'], errors='coerce')
                            D['LD'] = pd.to_datetime(D['LD'], origin='1899-12-30', unit='D', errors='coerce')
                            D['LD'] =  D['LD'].astype(str)
                            D[['Lyear', 'Lmonth', 'Lday']] = D['LD'].str.split('-', expand = True)
                        except:
                            pass
                        try:  
                            E['LD'] = pd.to_datetime(E['LD'],format='%d %m %Y', errors='coerce')
                            E['LD'] =  E['LD'].astype(str)
                            E[['Lyear', 'Lmonth', 'Lday']] = E['LD'].str.split('-', expand = True)
                        except:
                            pass
                        df = pd.concat([A,B,D,E])
                       
                        
                        #BRINGING BACK THE / IN DATES
                       
                 
                        df['RD'] = df['RD'].astype(str)
                      
                        df['TI'] = df['TI'].astype(str)
                        df['TO'] = df['TO'].astype(str)
                       
                        df['DD'] = df['DD'].astype(str)
                        df['LD'] = df['LD'].astype(str)
                        
            #        
                        
                        df['RD'] = df['RD'].str.replace('NaT', '',regex=True)
                      
                        df['TI'] = df['TI'].str.replace('NaT', '',regex=True)
                        df['TO'] = df['TO'].str.replace('NaT', '',regex=True)
                    
                        df['DD'] = df['DD'].str.replace('NaT', '',regex=True)
                        df['LD'] = df['LD'].str.replace('NaT', '',regex=True)
                       
                        
                        #SORTING THE TI YEARS
                        df[['Tiyear', 'Tiday']] =df[['Tiyear','Tiday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Tiyear'] = df['Tiyear'].fillna(994)
                        a = df[df['Tiyear']>31].copy()
                        b = df[df['Tiyear']<32].copy()
                        b = b.rename(columns={'Tiyear': 'Tiday2', 'Tiday': 'Tiyear'})
                        b = b.rename(columns={'Tiday2': 'Tiday'})
                        df = pd.concat([a,b])
                        

                        # #SORTING THE RETURN VISIT DATE YEARS
                        df[['Rday', 'Ryear']] = df[['Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        df['Ryear'] = df['Ryear'].fillna(994)
                        a = df[df['Ryear']>31].copy()
                        b = df[df['Ryear']<32].copy()
                        b = b.rename(columns={'Ryear': 'Rday2', 'Rday': 'Ryear'})
                        b = b.rename(columns={'Rday2': 'Rday'})
                        df = pd.concat([a,b])
                        
                            #SORTING THE TRANSFER OUT DATE YEAR
                        df[['Tday', 'Tyear']] = df[['Tday', 'Tyear']].apply(pd.to_numeric, errors='coerce')
                        df['Tyear'] = df['Tyear'].fillna(994)
                        a = df[df['Tyear']>31].copy()
                        b = df[df['Tyear']<32].copy()
                        b = b.rename(columns={'Tyear': 'Tday2', 'Tday': 'Tyear'})
                        b = b.rename(columns={'Tday2': 'Tday'})
                        df = pd.concat([a,b])         
                 
                        
                        #SORTING THE ART START YEARS
                        df[['Dyear', 'Dmonth', 'Dday']] =df[['Dyear', 'Dmonth', 'Dday']].apply(pd.to_numeric, errors = 'coerce')
                        df['Dyear'] = df['Dyear'].fillna(994)
                        a = df[df['Dyear']>31].copy()
                        b = df[df['Dyear']<32].copy()
                        b = b.rename(columns={'Dyear': 'Dday2', 'Dday': 'Dyear'})
                        b = b.rename(columns={'Dday2': 'Dday'})
                        df = pd.concat([a,b])
    
                        
                  
                        # #SORTING THE LAST ENCOUNTER
                        df[['Lday', 'Lyear']] = df[['Lday', 'Lyear']].apply(pd.to_numeric, errors='coerce')
                        df['Lyear'] = df['Lyear'].fillna(994)
                        a = df[df['Lyear']>31].copy()
                        b = df[df['Lyear']<32].copy()
                        b = b.rename(columns={'Lyear': 'Lday2', 'Lday': 'Lyear'})
                        b = b.rename(columns={'Lday2': 'Lday'})
                        df = pd.concat([a,b])
                        df = df.copy()
                        dfc = df.shape[0]
                   
                    
          
                        #COPY FOR ONE YEAR BEFORE GETTING POT CURR
                        def ager1(b):
                                if b < 1 :
                                    return '<1'
                                elif b < 5:
                                    return '1-4'
                                elif b < 10:
                                    return '5-9'
                                elif b < 15:
                                    return '10-14'
                                elif b < 20:
                                    return '15-19'
                                elif b < 25:
                                    return '20-24'
                                elif b < 30:
                                    return '25-29'
                                elif b < 35:
                                    return '30-34'
                                elif b < 40:
                                    return '35-39'
                                elif b < 45:
                                    return '40-44'
                                elif b < 50:
                                    return '45-49'
                                elif b < 55:
                                    return '50-54'
                                elif b < 60:
                                    return '55-59'
                                elif b < 65:
                                    return '60-64'
                                else:
                                    return '65+'

                        def ager2(a):
                                if a <5:
                                    return '0-4'
                                elif a < 10:
                                    return '5-9'
                                elif a < 15:
                                    return '10-14'
                                elif a < 20:
                                    return '15-19'
                                elif a < 25:
                                    return '20-24'
                                elif a < 30:
                                    return '25-29'
                                elif a < 40:
                                    return '30-39'
                                elif a < 50:
                                    return '40-49'
                                elif a >49:
                                    return '50+'
                        df['AG']  = pd.to_numeric(df['AG'], errors='coerce')
                        df['AGE BANDP'] = df['AG'].apply(ager2)
                        
                        df['AGE_BANDD'] = df['AG'].apply(ager2)
                        daty = {
                             'AGE BANDD': ['0-4', '5-9', '10-14', '15-19', '20-24',
                                          '25-29', '30-39', '40-49', '50+']
                        }
                        dfcompd = pd.DataFrame(daty)

                        datyx = {
                             'AGE BANDP': ['<1','1-4', '5-9', '10-14', '15-19', '20-24',
                                          '25-29', '30-34','35-39','40-44','45-49', '50-54','55-59','60-64','65+']
                        }
                        dfcompp = pd.DataFrame(datyx)
                
                        
        ################################ PARAMETERS ############################################
                        #Q1 parameters
                        byear = 2024 #one year ago
                        cyear = 2025  #curr year
                        cyp = 2026 # a year after
                        cyp1 = cyp +1
                        cmonth = 12 #last month of the qtr
                        cml = 11 # a month before
                        cmp = 13 # a month after
                        cday  = 4 #starting day
                        cdp = 5   # a day after
                        cdm = 3 #a day before
                        qmonths = [10,11,12] # months of the qtr
                        lmonth = 9 # last qtr month
        
                        lyear = 2025 # last qtr 
                        lday = 3  #last qtr day
                        ldp = 4 # a day after
                        ldm = 2 # a day before
                        vyeara = 2025 # current vl year
                        vyearb = 2024 # last vl year
                        vmonth = 0 # last vl month
                        vmm = 1  # a month after
                        oyear = 2024  #making one year
                        vayear = 2025 #for art start date in vl
                        vamonth = 7 # for art start date in vl
                        fmonth =  10 # first month of this qtr
                        tmonths = [7,8,9] # months of last qtr
                        ltmonth = 10 #first month of this qtr not eligible for DSD
                        m1 = 10
                        m2 = 11
                        m3 = 12 ####months of this qtr
                        tpy = 2025
                        tpm = 4
                        tpl = 3  ###MAKING 6 months on ART

                        #6months ago months
                        q6months = [4,5,6]
                        q6m = 4
                        q6year = 2025
                        #####
                        q1yr = 2025
                        year24 = 2023 # FOR 24 months cohort
                        
        
##################################################################################################################################
                        #POTENTIAL TXCUR ALTER... 
                        df[['Rmonth', 'Rday', 'Ryear']] = df[['Rmonth', 'Rday', 'Ryear']].apply(pd.to_numeric, errors='coerce')
                        df25 = df[df['Ryear']>lyear].copy()
                        df24 = df[df['Ryear'] == lyear].copy()
                        df24[['Rmonth', 'Rday']] = df24[['Rmonth', 'Rday']].apply(pd.to_numeric, errors='coerce')
                        df24 = df24[((df24['Rmonth']>lmonth) | ((df24['Rmonth']==lmonth) & (df24['Rday']>ldm)))].copy()
                        df = pd.concat([df25, df24]).copy()
                        
                        df = df.copy()
        
                        #REMOVE TO of the last reporting month
                        df[ 'Tyear'] = pd.to_numeric(df['Tyear'], errors='coerce')
                        dfto = df[df['Tyear']!=994].copy() #HAVE TOs
                        dfnot = df[df['Tyear'] == 994].copy() #NO TO
        
                       #REMOVE THE TO
                        dfto[['Ryear', 'Rmonth']] = dfto[['Ryear', 'Rmonth']].apply(pd.to_numeric, errors='coerce')
                        dftoy = dfto[((dfto['Ryear']!=lyear) |((dfto['Ryear']==lyear) & (dfto['Rmonth']>lmonth)))].copy() #OTHERS WOULD BE FALSE TOs, even those made last Q since they were brought as false if their RRDs were this year
                        
                        dftox = dfto[((dfto['Ryear']==lyear) & (dfto['Rmonth']==lmonth))].copy() #CLIENTS WITH RD OF REPORTING MONTH, DIDN'T RETURN BUT WERE TO LATER
                        dftox[['Tyear', 'Tmonth']] = dftox[['Tyear', 'Tmonth']].apply(pd.to_numeric, errors='coerce')
                        dftox = dftox[((dftox['Tyear']==lyear) & (dftox['Tmonth']>lmonth))].copy()
        
                        df = pd.concat([dftoy,dfnot])
                        df = df.copy()
                        if dftox.shape[0]>0:
                            df = pd.concat([df,dftox])
                        else:
                            df =df.copy()
                        #REMOVE the dead of the reporting month
                        df[ 'Dyear'] = pd.to_numeric(df['Dyear'], errors='coerce')
                        dfdd = df[df['Dyear']!=994].copy()
                        dfnot = df[df['Dyear'] == 994].copy()
                        #THOSE WHO DIED BEFORE FIRST MONTH OF THE Q
                        dfdd[['Dyear', 'Dmonth']] = dfdd[['Dyear', 'Dmonth']].apply(pd.to_numeric, errors='coerce')
                        dfdd = dfdd[((dfdd['Dyear']>lyear) |((dfdd['Dyear']==lyear) & (dfdd['Dmonth']>lmonth)))].copy() #DOESN'T MAKE SENSE
                       
                        df = pd.concat([ dfdd,dfnot])
            
                        df[['Lyear', 'Lmonth']]  = df[['Lyear', 'Lmonth']].apply(pd.to_numeric, errors='coerce')
                        dfpt = df[((df['Lyear']==cyear) & (df['Lmonth'].isin(qmonths)))].copy()

                        df = dfpt.copy()
                        screened = df.copy()
                        st.write(dfpt.shape[0])

                        htn = df[['AGE BANDP', 'HTN', 'ART', 'GD']].copy()

                        dm = df[['AGE BANDP', 'DM', 'ART', 'GD']].copy()
                        mh = df[['AGE BANDP', 'MH', 'ART', 'GD']].copy()

                        htn['HTN'] = htn['HTN'].astype(str)
                        # htn['HTN']  = htn['HTN'].fillna('Normal Blood pressure')
                        # htn['HTN'] = htn['HTN'].str.replace('nan', 'Normal Blood pressure') 
                        # htn['HTN'] = htn['HTN'].str.replace('NaT', 'Normal Blood pressure')

                        dm['DM'] = dm['DM'].astype(str)

                        h1map = {
                            'Known Hypertensive, not controlled and on lifestyle modification and Medication':'Known',
                            'Newly Diagnosed and on lifestyle modification': 'Newly',
                            'Known Hypertensive but not on Treatment':'Known',
                            'Known Hypertensive, not controlled and on lifestyle modification':'Known',
                            'Known Hypertensive and Controlled and on lifestyle modification and Medication':'Known',
                            'Known Hypertensive and Controlled and on lifestyle modification':'Known',
                            'Newly Diagnosed and on lifestyle modification and Medication':'Newly',
                            'Hypertensive patient referred for further management': 'Refered'
                        }

                        h2map = {
                            'Known Hypertensive, not controlled and on lifestyle modification and Medication':'LFT',
                            'Newly Diagnosed and on lifestyle modification': 'LF',
                            'Known Hypertensive, not controlled and on lifestyle modification':'LF',
                            'Known Hypertensive and Controlled and on lifestyle modification and Medication':'LFT',
                            'Known Hypertensive and Controlled and on lifestyle modification':'LF',
                            'Newly Diagnosed and on lifestyle modification and Medication':'LFT'
                        }

                        h3map = {
                            'Known Hypertensive, not controlled and on lifestyle modification and Medication':'NCT',
                            'Known Hypertensive, not controlled and on lifestyle modification':'NCT',
                            'Known Hypertensive and Controlled and on lifestyle modification and Medication':'CT',
                            'Known Hypertensive and Controlled and on lifestyle modification':'CT'
                        }

                        htn['SCR'] = htn['HTN'].map(h1map)
                        htn['RX'] = htn['HTN'].map(h2map)
                        htn['CTR'] = htn['HTN'].map(h3map)

                        nscr = htn[htn['SCR'].isin(['Known', 'Newly', 'Refered'])].copy()
                        kscr = htn[htn['RX'].isin(['LF', 'LFT'])].copy()
                        cscr = htn[htn['CTR'].isin(['NCT','CT'])].copy()


    
#DM DATA   3333333333333333333333333333333333333333333333333333
                        d1map = {
                            'Known Diabetic and not controlled on medication and lifestyle modification':'Known',
                            'Newly Diagnosed and on lifestyle modification': 'Newly',
                            'Newly Diagnosed and on lifestyle modification and Medication':'Newly',
                            'Diabetic patient referred for further management':'Refered',
                            'Known Diabetic and controlled on medication and lifestyle modification':'Known'
                        }

                        d2map = {
                            'Known Diabetic and not controlled on medication and lifestyle modification':'LFT',
                            'Newly Diagnosed and on lifestyle modification': 'LF',
                            'Newly Diagnosed and on lifestyle modification and Medication':'LFT',
                            'Known Diabetic and controlled on medication and lifestyle modification':'LFT',
                        }

                        d3map = {
                             'Known Diabetic and not controlled on medication and lifestyle modification':'NCT',
                            'Known Diabetic and controlled on medication and lifestyle modification':'CT',
                        }

                        dm['SCR'] = dm['DM'].map(d1map)
                        dm['RX'] = dm['DM'].map(d2map)
                        dm['CTR'] = dm['DM'].map(d3map)

                        ndscr = dm[dm['SCR'].isin(['Known', 'Newly', 'Refered'])].copy()
                        ndkcr = dm[dm['RX'].isin(['LF', 'LFT'])].copy()
                        ndccr = dm[dm['CTR'].isin(['NCT','CT'])].copy()

#MH DATA   3333333333333333333333333333333333333333333333333333
                        m1map = {
                            'Known mental illness on Psychotherapy and Medication':'Known',
                            'Has signs and symptoms with Low/moderate suicide Risk': 'Newly',
                            'Known mental illness on Psychotherapy only':'Known',
                        }

                        m2map = {
                            'Known mental illness on Psychotherapy and Medication':'LFT',
                            'Known mental illness on Psychotherapy only':'LF',
                        }

                        # m3map = {
                        #         'Known Diabetic and not controlled on medication and lifestyle modification':'NCT',
                        #     'Known Diabetic and controlled on medication and lifestyle modification':'NCT',
                        # }
                        mh['SCR'] = mh['MH'].map(m1map)
                        mh['RX'] = mh['MH'].map(m2map)

                        mdscr = mh[mh['SCR'].isin(['Known', 'Newly', 'Refered'])].copy()
                        mdkcr = mh[mh['RX'].isin(['LF', 'LFT'])].copy()
                       
  
    if st.session_state.readery:# and st.session_state.dfw:
        
            # Create an in-memory BytesIO buffer
            output = io.BytesIO()
            mapperp = {'0-4':1, '5-9':2, '10-14':3, '15-19':4, '20-24':5, '25-29':6, '30-39':7, '40-49':8,'50+':9}
            
            mapperd = {'<1':1,'1-4':2, '5-9':3, '10-14':4, '15-19':5, '20-24':6, '25-29':7, '30-34':8, '35-39':9,
                  '40-44':10, '45-49':11, '50-54':12, '55-59':13, '60-64':14, '65+':15}                
            with pd.ExcelWriter(output, engine="xlsxwriter") as writer:

                        hc07 = screened.groupby(['AGE BANDP', 'GD']).size().unstack(fill_value=0)
                        hc07 = hc07.reindex(columns=['M', 'F'], fill_value=0)
                        hc07 = hc07.reset_index()
                        hc07 = pd.merge(dfcompp, hc07, on='AGE BANDP', how='left')
                        hc07['M'] = hc07['M'].fillna(0)
                        hc07['F'] = hc07['F'].fillna(0) 
                        # hc07['AGE BAND'] = hc07['AGE BAND'].astype(str)
                        hc07['SORT'] = hc07['AGE BANDP'].map(mapperp)
                        hc07['SORT'] = pd.to_numeric(hc07['SORT'], errors='coerce')
                        hc07 = hc07.sort_values(by ='SORT')
                        hc07 = hc07.drop(columns='SORT')
                        hc07.to_excel(writer, sheet_name="SCREENED", index=False)
###################################################################################################
                        hc06 = nscr.groupby(['SCR', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc06 = hc06.reindex(columns=['M', 'F'], fill_value=0)
                        hc06 = hc06.reset_index()
                        # hc06 = pd.merge(dfcompp, hc06, on='AGE BANDP', how='left')
                        hc06['M'] = hc06['M'].fillna(0)
                        hc06['F'] = hc06['F'].fillna(0) 
                        mapperc = {'Newly':1, 'Known':2, 'Refered':3}
                        # hc06['AGE BAND'] = hc06['AGE BAND'].astype(str)
                        hc06['SORTX'] = hc06['SCR'].map(mapperc)
                        hc06['SORTX'] = pd.to_numeric(hc06['SORTX'], errors='coerce')

                        hc06['SORT'] = hc06['AGE BANDP'].map(mapperp)
                        hc06['SORT'] = pd.to_numeric(hc06['SORT'], errors='coerce')
                        hc06 = hc06.sort_values(by =['SORTX','SORT'])
                        hc06 = hc06.drop(columns=['SORTX','SORT'])
                        hc06.to_excel(writer, sheet_name="HNEWLY", index=False)

                        #################################################################

                        hc08 = kscr.groupby(['RX', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc08 = hc08.reindex(columns=['M', 'F'], fill_value=0)
                        hc08 = hc08.reset_index()

                       
                        hc08['M'] = hc08['M'].fillna(0)
                        hc08['F'] = hc08['F'].fillna(0)
                        hc08['SORT'] = hc08['AGE BANDP'].map(mapperp)
                        mapperc = {'LF':1, 'LFT':2}
                        hc08['SORTX'] = hc08['RX'].map(mapperc)
                        hc08['SORTX'] = pd.to_numeric(hc08['SORTX'], errors='coerce')
                        
                        hc08['SORT'] = pd.to_numeric(hc08['SORT'], errors='coerce')
                        hc08 = hc08.sort_values(by =['SORTX','SORT'])
                        hc08 = hc08.drop(columns=['SORTX','SORT'])
                        hc08.to_excel(writer, sheet_name="HKNOWN", index=False)

    #######################################################################################

                        hc09 = cscr.groupby(['CTR', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc09 = hc09.reindex(columns=['M', 'F'], fill_value=0)
                        hc09 = hc09.reset_index()
                       
                        hc09['M'] = hc09['M'].fillna(0)
                        hc09['F'] = hc09['F'].fillna(0) 
                        hc09['SORT'] = hc09['AGE BANDP'].map(mapperp)
                        mapperc = {'CT':1, 'NCT':2}
                        hc09['SORTX'] = hc09['CTR'].map(mapperc)
                        hc09['SORTX'] = pd.to_numeric(hc09['SORTX'], errors='coerce')

                        hc09['SORT'] = pd.to_numeric(hc09['SORT'], errors='coerce')
                        hc09 = hc09.sort_values(by =['SORTX','SORT'])
                        hc09 = hc09.drop(columns=['SORTX','SORT'])
                        hc09.to_excel(writer, sheet_name="HCTRL", index=False)

##########################################3DIABETES
                        hc06 = ndscr.groupby(['SCR', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc06 = hc06.reindex(columns=['M', 'F'], fill_value=0)
                        hc06 = hc06.reset_index()
                        # hc06 = pd.merge(dfcompp, hc06, on='AGE BANDP', how='left')
                        hc06['M'] = hc06['M'].fillna(0)
                        hc06['F'] = hc06['F'].fillna(0) 
                        mapperc = {'Newly':1, 'Known':2, 'Refered':3}
                        # hc06['AGE BAND'] = hc06['AGE BAND'].astype(str)
                        hc06['SORTX'] = hc06['SCR'].map(mapperc)
                        hc06['SORTX'] = pd.to_numeric(hc06['SORTX'], errors='coerce')

                        hc06['SORT'] = hc06['AGE BANDP'].map(mapperp)
                        hc06['SORT'] = pd.to_numeric(hc06['SORT'], errors='coerce')
                        hc06 = hc06.sort_values(by =['SORTX','SORT'])
                        hc06 = hc06.drop(columns=['SORTX','SORT'])
                        hc06.to_excel(writer, sheet_name="DNEWLY", index=False)

                        #################################################################

                        hc08 = ndscr.groupby(['RX', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc08 = hc08.reindex(columns=['M', 'F'], fill_value=0)
                        hc08 = hc08.reset_index()

                       
                        hc08['M'] = hc08['M'].fillna(0)
                        hc08['F'] = hc08['F'].fillna(0)
                        hc08['SORT'] = hc08['AGE BANDP'].map(mapperp)
                        mapperc = {'LF':1, 'LFT':2}
                        hc08['SORTX'] = hc08['RX'].map(mapperc)
                        hc08['SORTX'] = pd.to_numeric(hc08['SORTX'], errors='coerce')
                        
                        hc08['SORT'] = pd.to_numeric(hc08['SORT'], errors='coerce')
                        hc08 = hc08.sort_values(by =['SORTX','SORT'])
                        hc08 = hc08.drop(columns=['SORTX','SORT'])
                        hc08.to_excel(writer, sheet_name="DKNOWN", index=False)

                        ####################################################################################

                        hc09 = ndccr.groupby(['CTR', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc09 = hc09.reindex(columns=['M', 'F'], fill_value=0)
                        hc09 = hc09.reset_index()
                       
                        hc09['M'] = hc09['M'].fillna(0)
                        hc09['F'] = hc09['F'].fillna(0) 
                        hc09['SORT'] = hc09['AGE BANDP'].map(mapperp)
                        mapperc = {'CT':1, 'NCT':2}
                        hc09['SORTX'] = hc09['CTR'].map(mapperc)
                        hc09['SORTX'] = pd.to_numeric(hc09['SORTX'], errors='coerce')

                        hc09['SORT'] = pd.to_numeric(hc09['SORT'], errors='coerce')
                        hc09 = hc09.sort_values(by =['SORTX','SORT'])
                        hc09 = hc09.drop(columns=['SORTX','SORT'])
                        hc09.to_excel(writer, sheet_name="DCTRL", index=False)
                        
##########################################MENTAL HEALTH
                        hc06 = mdscr.groupby(['SCR', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc06 = hc06.reindex(columns=['M', 'F'], fill_value=0)
                        hc06 = hc06.reset_index()
                        # hc06 = pd.merge(dfcompp, hc06, on='AGE BANDP', how='left')
                        hc06['M'] = hc06['M'].fillna(0)
                        hc06['F'] = hc06['F'].fillna(0) 
                        mapperc = {'Newly':1, 'Known':2, 'Refered':3}
                        # hc06['AGE BAND'] = hc06['AGE BAND'].astype(str)
                        hc06['SORTX'] = hc06['SCR'].map(mapperc)
                        hc06['SORTX'] = pd.to_numeric(hc06['SORTX'], errors='coerce')

                        hc06['SORT'] = hc06['AGE BANDP'].map(mapperp)
                        hc06['SORT'] = pd.to_numeric(hc06['SORT'], errors='coerce')
                        hc06 = hc06.sort_values(by =['SORTX','SORT'])
                        hc06 = hc06.drop(columns=['SORTX','SORT'])
                        hc06.to_excel(writer, sheet_name="MNEWLY", index=False)

                        #################################################################

                        hc08 = mdscr.groupby(['RX', 'AGE BANDP', 'GD']).size().unstack('GD', fill_value=0)
                        hc08 = hc08.reindex(columns=['M', 'F'], fill_value=0)
                        hc08 = hc08.reset_index()

                       
                        hc08['M'] = hc08['M'].fillna(0)
                        hc08['F'] = hc08['F'].fillna(0)
                        hc08['SORT'] = hc08['AGE BANDP'].map(mapperp)
                        mapperc = {'LF':1, 'LFT':2}
                        hc08['SORTX'] = hc08['RX'].map(mapperc)
                        hc08['SORTX'] = pd.to_numeric(hc08['SORTX'], errors='coerce')
                        
                        hc08['SORT'] = pd.to_numeric(hc08['SORT'], errors='coerce')
                        hc08 = hc08.sort_values(by =['SORTX','SORT'])
                        hc08 = hc08.drop(columns=['SORTX','SORT'])
                        hc08.to_excel(writer, sheet_name="MKNOWN", index=False)
 


    
            output.seek(0)    
                            # Provide one combined download button
            facility = st.text_input('FACILITY NAME')
            if not facility:
                             st.stop()
            else:
                            st.download_button(
                                label="📥 DOWNLOAD PIVOT TABLES",
                                data=output,
                                file_name=f"{facility}_NCDS_DHIS2.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
     

                            st.success('**CREATED BY Dr. LUMINSA DESIRE**')
                        
pages = {
    "READER:": [
        st.Page(extract, title="EMR EXTRACT READER"),
    ],
    "QUARTERLY:":[
        st.Page("qtr.py", title="DASHBOARD"),
        ]
}

pg = st.navigation(pages)
pg.run()
                                
    